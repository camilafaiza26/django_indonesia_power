from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from master.models import *
from .models import *
from .forms import *
import datetime
from django.utils import timezone
from django.db.models import Sum
from pulp import *
import pulp as p
import pandas as pd
import numpy as np
import simplejson as json
from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.db.models import Q


# Create your views here.
def loginView(request):
    print("Masuk1",request.method)
    #print(request.GET)
    print("Masuk1a",request.POST)
    context = {
        'page_title':'LOGIN',
    }
    user = None

    if request.method == "GET":
        if request.user.is_authenticated:
            # logika untuk user
            print("Masuk2", request.method)
            #print(request.GET)
            print("masuk2a",request.GET)
            return redirect('index')
        else:
            # logika untuk anonymous
            #return render(request, 'login.html', context)
            print("Masuk3", request.method)
            #print(request.GET)
            print("masuk3a",request.GET)
            return redirect('login')


    if request.method == "POST":
        print("Masuk4", request.method)
        #print(request.GET)
        print("masuk4a",request.POST)
        username_login = request.POST['username']
        password_login = request.POST['password']

        user = authenticate(request, username=username_login, password=password_login)

        if user is not None:
            print("Masuk5", request.method)
            #print(request.GET)
            print("masuk5a",request.POST)
            login(request, user)
            return redirect('index')
        else:
            print("Masuk6", request.method)
            #print(request.GET)
            print("masuk6a",request.POST)
            return redirect('login')

    return redirect('login')

def index(request):
    print(request.method)
    #print(request.GET)
    print(request.POST)
    if request.user.is_authenticated:
        '''if request.method == "POST":
            if request.POST["logout"] == "":
                logout(request)
                return redirect('login')'''
        if 'logout' in request.POST:
            logoutnya = request.POST['logout']
            if logoutnya == "":
                logout(request)
                return redirect('login')
        if 'sort' in request.POST:
            sort = request.POST['sort']
            if sort != None:
                #queryTableMasterSolver = FactSolverKontrakPemasok.objects.all().order_by('-id_kontrakpemasok')
                #queryTableMasterSolver = FactSolverKontrakPemasok.objects.all()
                queryPLTU = PLTU.objects.all()
                queryPemasok = Pemasok.objects.all()
                #querySumberTambang = Sumbertambang.objects.all()
                #queryWaktu = Waktu.objects.all()
                queryWaktuDistinctTriwulan = Waktu.objects.filter(~Q(tahun=2019)).values('triwulan').distinct()
                queryWaktuDistinctTahun = Waktu.objects.all().values('tahun').distinct()
                querySumberTambangDistinct = Sumbertambang.objects.all().values('nama_sumber_tambang').distinct()

                sortpltu = request.POST['sortpltu']
                query_get_id_pltu = PLTU.objects.get(nama_pltu=sortpltu)
                id_pltunya = query_get_id_pltu.pltu_id
                querypltufilter = FactSolverKontrakPemasok.objects.filter(pltu = id_pltunya)
                #mastersolverform = MasterSolverForm()
                #contact_form = ContactForm()
                context = {
                    'judul':'Solver',
                    'kontributor':'Admin',
                    'nav': [
                        ['/','Home'],
                        ['/blog/news','News'],
                        ['/about','About'],
                    ],
                    'app_css':'blog/css/styles.css',
                    #'mastersolverform':mastersolverform,
                    #'contact_form':contact_form,
                    #'dataMasterSolver':queryTableMasterSolver,
                    'queryPLTU':queryPLTU,
                    'queryPemasok':queryPemasok,
                    #'querySumberTambang':querySumberTambang,
                    #'queryWaktu':queryWaktu,
                    'queryWaktuDistinctTriwulan':queryWaktuDistinctTriwulan,
                    'queryWaktuDistinctTahun':queryWaktuDistinctTahun,
                    'querypltufilter':querypltufilter,
                    'querySumberTambangDistinct':querySumberTambangDistinct,

                }
                return render(request,'about/index.html', context)

        else:
            #queryTableMasterSolver = FactSolverKontrakPemasok.objects.all().order_by('-id_kontrakpemasok')
            queryTableMasterSolver = FactSolverKontrakPemasok.objects.all()
            queryPLTU = PLTU.objects.all()
            queryPemasok = Pemasok.objects.all()
            #querySumberTambang = Sumbertambang.objects.all()
            #queryWaktu = Waktu.objects.all()
            queryWaktuDistinctTriwulan = Waktu.objects.filter(~Q(tahun=2019)).values('triwulan').distinct()
            queryWaktuDistinctTahun = Waktu.objects.all().values('tahun').distinct()
            querySumberTambangDistinct = Sumbertambang.objects.all().values('nama_sumber_tambang').distinct()

            #mastersolverform = MasterSolverForm()
            #contact_form = ContactForm()
            context = {
                'judul':'Solver',
                'kontributor':'Admin',
                'nav': [
                    ['/','Home'],
                    ['/blog/news','News'],
                    ['/about','About'],
                ],
                'app_css':'blog/css/styles.css',
                #'mastersolverform':mastersolverform,
                #'contact_form':contact_form,
                'dataMasterSolver':queryTableMasterSolver,
                'queryPLTU':queryPLTU,
                'queryPemasok':queryPemasok,
                #'querySumberTambang':querySumberTambang,
                #'queryWaktu':queryWaktu,
                'queryWaktuDistinctTriwulan':queryWaktuDistinctTriwulan,
                'queryWaktuDistinctTahun':queryWaktuDistinctTahun,
                'querySumberTambangDistinct':querySumberTambangDistinct,


            }

            return render(request,'about/index.html', context)
    else:
        #return render(request, 'login.html')
        return redirect('login')


def inputsolver(request):
    print("POSTinputsolver",request.POST.items())
    print("POSTinputsolver",request.POST)
    print("POSTinputsolver",request.POST.__dict__)
    #print(timezone.localtime(timezone.now()))
    if request.user.is_authenticated:
        '''if request.method == "POST":
            if request.POST["logout"] == "":
                logout(request)
                return redirect('login')'''
        if 'logout' in request.POST:
            logoutnya = request.POST['logout']
            if logoutnya == "":
                logout(request)
                return redirect('login')
        if 'pilihpltu' in request.POST:
            pilihpltu = request.POST['pilihpltu']
            if pilihpltu != None:
                #queryTableMasterSolver = FactSolverKontrakPemasok.objects.all()
                queryPLTU = PLTU.objects.all()
                #queryPemasok = Pemasok.objects.all()
                #querySumberTambang = Sumbertambang.objects.all()
                #queryWaktu = Waktu.objects.all()
                #queryWaktuDistinctTriwulan = Waktu.objects.all().values('triwulan').distinct()
                #queryWaktuDistinctTahun = Waktu.objects.all().values('tahun').distinct()

                pltu = request.POST['pltu']
                query_get_id_pltu = PLTU.objects.get(nama_pltu=pltu)
                id_pltunya = query_get_id_pltu.pltu_id
                queryPemasokDistinct = FactSolverKontrakPemasok.objects.filter(pltu_id=id_pltunya).values('pemasok__nama_pemasok').distinct()

                querySetRencana = FactSolverRencana.objects.filter(pltu_id=id_pltunya)

                queryAmbilRangePasokan = FactSolverRangePasokanDiinginkan.objects.filter(pltu_id=id_pltunya)
                #print(queryAmbilRangePasokan.minimal_volume)
                #print("pilihpltu",request.POST)
                total_min = 0
                total_max = 0
                for i in queryAmbilRangePasokan:
                    a = i.minimal_total_perbulan
                    b = i.maksimal_total_perbulan
                    total_min = total_min + a
                    total_max = total_max + b

                context2 = {
                    'judul':'Solver',
                    'kontributor':'Admin',
                    'nav': [
                        ['/','Home'],
                        ['/blog/news','News'],
                        ['/about','About'],
                    ],
                    'app_css':'blog/css/styles.css',

                    #'dataMasterSolver':queryTableMasterSolver,
                    'queryPLTU':queryPLTU,
                    #'queryPemasok':queryPemasok,
                    #'querySumberTambang':querySumberTambang,
                    #'queryWaktu':queryWaktu,
                    #'queryWaktuDistinctTriwulan':queryWaktuDistinctTriwulan,
                    #'queryWaktuDistinctTahun':queryWaktuDistinctTahun,
                    #'queryPemasokDistinct':queryPemasokDistinct,
                    'pltuterpilih':pltu,
                    'querySetRencana':querySetRencana,
                    'queryAmbilRangePasokan':queryAmbilRangePasokan,
                    'total_min':total_min,
                    'total_max':total_max



                }

                return render(request,'about/inputsolver.html', context2)

        if 'editrencanasolver' in request.POST:
            editrencanasolver = request.POST['editrencanasolver']
            if editrencanasolver != None:
                #queryTableMasterSolver = FactSolverKontrakPemasok.objects.all()
                queryPLTU = PLTU.objects.all()
                #queryPemasok = Pemasok.objects.all()
                #querySumberTambang = Sumbertambang.objects.all()
                #queryWaktu = Waktu.objects.all()
                #queryWaktuDistinctTriwulan = Waktu.objects.all().values('triwulan').distinct()
                #queryWaktuDistinctTahun = Waktu.objects.all().values('tahun').distinct()

                pltu = request.POST['pltu']
                query_get_id_pltu = PLTU.objects.get(nama_pltu=pltu)
                id_pltunya = query_get_id_pltu.pltu_id
                #queryPemasokDistinct = FactSolverKontrakPemasok.objects.filter(pltu_id=id_pltunya).values('pemasok__nama_pemasok').distinct()

                querySetRencana = FactSolverRencana.objects.filter(pltu_id=id_pltunya)

                context2 = {
                    'judul':'Solver',
                    'kontributor':'Admin',
                    'nav': [
                        ['/','Home'],
                        ['/blog/news','News'],
                        ['/about','About'],
                    ],
                    'app_css':'blog/css/styles.css',

                    #'dataMasterSolver':queryTableMasterSolver,
                    'queryPLTU':queryPLTU,
                    #'queryPemasok':queryPemasok,
                    #'querySumberTambang':querySumberTambang,
                    #'queryWaktu':queryWaktu,
                    #'queryWaktuDistinctTriwulan':queryWaktuDistinctTriwulan,
                    #'queryWaktuDistinctTahun':queryWaktuDistinctTahun,
                    #'queryPemasokDistinct':queryPemasokDistinct,
                    'pltuterpilih':pltu,
                    'querySetRencana':querySetRencana

                }

                return render(request,'about/editrencanakebutuhan.html', context2)
                #return redirect('updaterencanakebutuhan')

        if 'editrangepasokan' in request.POST:
            editrangepasokan = request.POST['editrangepasokan']
            if editrangepasokan != None:
                #queryTableMasterSolver = FactSolverKontrakPemasok.objects.all()
                queryPLTU = PLTU.objects.all()
                #queryPemasok = Pemasok.objects.all()
                #querySumberTambang = Sumbertambang.objects.all()
                #queryWaktu = Waktu.objects.all()
                #queryWaktuDistinctTriwulan = Waktu.objects.all().values('triwulan').distinct()
                #queryWaktuDistinctTahun = Waktu.objects.all().values('tahun').distinct()

                pltu = request.POST['pltu']
                query_get_id_pltu = PLTU.objects.get(nama_pltu=pltu)
                id_pltunya = query_get_id_pltu.pltu_id
                #queryPemasokDistinct = FactSolverKontrakPemasok.objects.filter(pltu_id=id_pltunya).values('pemasok__nama_pemasok').distinct()

                querySetRencana = FactSolverRencana.objects.filter(pltu_id=id_pltunya)

                queryAmbilRangePasokan = FactSolverRangePasokanDiinginkan.objects.filter(pltu_id=id_pltunya)

                context2 = {
                    'judul':'Solver',
                    'kontributor':'Admin',
                    'nav': [
                        ['/','Home'],
                        ['/blog/news','News'],
                        ['/about','About'],
                    ],
                    'app_css':'blog/css/styles.css',

                    #'dataMasterSolver':queryTableMasterSolver,
                    'queryPLTU':queryPLTU,
                    #'queryPemasok':queryPemasok,
                    #'querySumberTambang':querySumberTambang,
                    #'queryWaktu':queryWaktu,
                    #'queryWaktuDistinctTriwulan':queryWaktuDistinctTriwulan,
                    #'queryWaktuDistinctTahun':queryWaktuDistinctTahun,
                    #'queryPemasokDistinct':queryPemasokDistinct,
                    'pltuterpilih':pltu,
                    'querySetRencana':querySetRencana,
                    'queryAmbilRangePasokan':queryAmbilRangePasokan,
                }

                return render(request,'about/editrangepasokan.html', context2)
                #return redirect('updaterencanakebutuhan')

        if 'simpanrencanasolver' in request.POST:
            simpanrencanasolver = request.POST['simpanrencanasolver']
            if simpanrencanasolver != None:

                request.POST._mutable = True

                pltu = request.POST['pltu']
                query_get_id_pltu = PLTU.objects.get(nama_pltu=pltu)
                a = query_get_id_pltu.pltu_id
                querypk = FactSolverRencana.objects.get(pltu_id=a)
                pknya = querypk.rencana_id
                request.POST['pltu']=a
                now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                request.POST['waktu']=now
                #b = request.POST['persentase_rencana_mrc_persen']
                #c = request.POST['persentase_rencana_lrc_persen']
                #d = request.POST['input_bb_mt']
                #e = request.POST['input_nilai_kalor_kcal_kg']
                #f = request.POST['input_total_sulphur']

                request.POST._mutable = False
                context4 = {
                    'judul':'Solver',
                    'kontributor':'Admin',
                    'nav': [
                        ['/','Home'],
                        ['/blog/news','News'],
                        ['/about','About'],
                    ],
                    'app_css':'blog/css/styles.css',


                }

                if request.method == 'POST':
                    instance = get_object_or_404(FactSolverRencana, pk=pknya)
                    form = ValidasiRencanaSolver(request.POST or None, instance=instance)
                    if form.is_valid():
                        form.save()
                else:
                    return redirect('inputsolver')

        if 'simpanrangepasokan' in request.POST:
            simpanrangepasokan = request.POST['simpanrangepasokan']
            if simpanrangepasokan != None:

                request.POST._mutable = True

                #queryTableMasterSolver = FactSolverKontrakPemasok.objects.all()
                queryPLTU = PLTU.objects.all()
                #queryPemasok = Pemasok.objects.all()
                #querySumberTambang = Sumbertambang.objects.all()
                #queryWaktu = Waktu.objects.all()
                #queryWaktuDistinctTriwulan = Waktu.objects.all().values('triwulan').distinct()
                #queryWaktuDistinctTahun = Waktu.objects.all().values('tahun').distinct()

                pltu = request.POST['pltu']
                query_get_id_pltu = PLTU.objects.get(nama_pltu=pltu)
                id_pltunya = query_get_id_pltu.pltu_id
                queryPemasokDistinct = FactSolverKontrakPemasok.objects.filter(pltu_id=id_pltunya).values('pemasok__nama_pemasok').distinct()
                count_querypemasokdistinct = FactSolverKontrakPemasok.objects.filter(pltu_id=id_pltunya).values('pemasok__nama_pemasok').distinct().count()
                request.POST['pltu'] = id_pltunya


                request.POST._mutable = False

                #print("AwalInputSolver",request.POST.items())
                #print("AwalInputSolver",request.POST)
                #print("AwalInputSolver",request.POST.__dict__)
                for x in range(count_querypemasokdistinct):
                    request.POST._mutable = True
                    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                    print(now)

                    pemasok = request.POST.getlist('pemasok')[x]
                    query_get_id_pemasok = Pemasok.objects.get(nama_pemasok=pemasok)
                    id_pemasoknya = query_get_id_pemasok.pemasok_id
                    request.POST.getlist('pemasok')[x] = id_pemasoknya
                    #a = request.POST.getlist('pemasok')[x]
                    a = id_pemasoknya
                    if request.POST.getlist('minimal_volume')[x] == "":
                        b=0
                    else:
                        b = float(request.POST.getlist('minimal_volume')[x])
                    if request.POST.getlist('minimal_kali_perbulan')[x] == "":
                        c=0
                    else:
                        c = int(request.POST.getlist('minimal_kali_perbulan')[x])
                    bc = b*c
                    #d = float(request.POST.getlist('minimal_total_perbulan')[x])
                    if request.POST.getlist('maksimal_volume')[x] == "":
                        e=0
                    else:
                        e = float(request.POST.getlist('maksimal_volume')[x])
                    if request.POST.getlist('maksimal_kali_perbulan')[x] == "":
                        f=0
                    else:
                        f = int(request.POST.getlist('maksimal_kali_perbulan')[x])
                    ef = e*f
                    #g = float(request.POST.getlist('maksimal_total_perbulan')[x])
                    barisnya = FactSolverRangePasokanDiinginkan.objects.get(pltu_id=id_pltunya, pemasok_id=id_pemasoknya)


                    request.POST._mutable = False

                    barisnya.waktu_0 = now
                    barisnya.minimal_volume = b
                    barisnya.minimal_kali_perbulan = c
                    barisnya.minimal_total_perbulan = bc
                    barisnya.maksimal_volume = e
                    barisnya.maksimal_kali_perbulan = f
                    barisnya.maksimal_total_perbulan = ef
                    barisnya.save()


                context3 = {
                    'judul':'Solver',
                    'kontributor':'Admin',
                    'nav': [
                        ['/','Home'],
                        ['/blog/news','News'],
                        ['/about','About'],
                    ],
                    'app_css':'blog/css/styles.css',

                    #'dataMasterSolver':queryTableMasterSolver,
                    'queryPLTU':queryPLTU,
                    #'queryPemasok':queryPemasok,
                    #'querySumberTambang':querySumberTambang,
                    #'queryWaktu':queryWaktu,
                    #'queryWaktuDistinctTriwulan':queryWaktuDistinctTriwulan,
                    #'queryWaktuDistinctTahun':queryWaktuDistinctTahun,
                    #'queryPemasokDistinct':queryPemasokDistinct,
                    #'pltuterpilih':pltu,



                }

                #print(request.POST.getlist('volume_min')[0])
                return render(request,'about/inputsolver.html', context3)

        if 'prosessolver' in request.POST:
            prosessolver = request.POST['prosessolver']
            if prosessolver != None:
                request.POST._mutable = True

                pltu = request.POST['pltu']
                query_get_id_pltu = PLTU.objects.get(nama_pltu=pltu)
                a = query_get_id_pltu.pltu_id
                querypk = FactSolverRencana.objects.get(pltu_id=a)
                pknya = querypk.rencana_id
                request.POST['pltu']=a
                now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                request.POST['waktu']=now

                statussulfur=""
                if 'statussulfur' in request.POST:
                    statussulfur = request.POST['statussulfur']

                request.POST._mutable = False
                context4 = {
                    'judul':'Solver',
                    'kontributor':'Admin',
                    'nav': [
                        ['/','Home'],
                        ['/blog/news','News'],
                        ['/about','About'],
                    ],
                    'app_css':'blog/css/styles.css',


                }

                if request.method == 'POST':

                    nama_pltu = pltu

                    print("\n ============= PROGRAM OPTIMASI BIAYA BATUBARA ================ \n\n")
                    print("\n Nama pltu yang diproses :  \n", nama_pltu)

                    list_sumber_tambang = list(FactSolverKontrakPemasok.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list('nama_sumber_tambang', flat=True))

                    # list id pemasok per baris
                    list_id_pemasok = list(FactSolverKontrakPemasok.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list('pemasok_id', flat=True))

                    # list id pemasok per pemasok (untuk table resume)
                    list_id_pemasok_distinct = list(FactSolverKontrakPemasok.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list('pemasok_id', flat=True).distinct())

                    # list id PLTU per baris
                    id_pltu = FactSolverKontrakPemasok.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list('pltu_id', flat=True).distinct()

                    # definisikan variable
                    # x1 maupun x2 tidak boleh bernilai negatif
                    list_volume = []
                    for i, a in enumerate(list_sumber_tambang):
                        globals()["x"+str(i+1)] = p.LpVariable("x" +
                                                            str(i+1), lowBound=0, cat='Float')
                        list_volume.append(globals()["x"+str(i+1)])

                    print("\n List X : \n", list_volume)
                    print("\n Banyaknya X adalah : ", len(list_volume))

                    total_volume = ""
                    for i in range(0, len(list_volume)):
                        total_volume += (list_volume[i])

                    list_pemasok_lontar = list(FactSolverKontrakPemasok.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list('pemasok_id__nama_pemasok', flat=True))

                    print("\n Isi List pemasok = \n", list_pemasok_lontar)

                    # list_pemasok = Data_solver.objects.values_list(
                    #   'nama_pemasok', flat=True)

                    input_rencana_batubara = FactSolverRencana.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list('input_bb_mt', flat=True).get()
                    print("\n Input rencana batubara  = ", input_rencana_batubara)

                    input_rencana_kalor = FactSolverRencana.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list(
                        'input_nilai_kalor_kcal_kg', flat=True).get()
                    print("\n Input rencana kalor  = ", input_rencana_kalor)

                    input_rencana_sulphur = FactSolverRencana.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list(
                        'input_total_sulphur', flat=True).get()
                    print("\n Isi input rencana Sulphur  = ", input_rencana_sulphur)

                    print("\n Isi List sumber tambang= \n", list_sumber_tambang)

                    persentase_batas_lrc = FactSolverRencana.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list(
                        'persentase_rencana_lrc_persen', flat=True).get()

                    persentase_batas_mrc = FactSolverRencana.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list(
                        'persentase_rencana_mrc_persen', flat=True).get()

                    # Inisialisasi nilai maksimal pasokan batubara
                    max_kali_per_bulan = list(FactSolverRangePasokanDiinginkan.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list('maksimal_kali_perbulan', flat=True))

                    max_volume = list(FactSolverRangePasokanDiinginkan.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list('maksimal_volume', flat=True))

                    print("\n Max Volume  = ", max_volume)

                    list_max_batubara = []

                    for i in range(0, len(max_kali_per_bulan)):
                        total_setahun = max_volume[i] * max_kali_per_bulan[i]
                        list_max_batubara.append(total_setahun)

                    print("\n Isi list Max batubara \n", list_max_batubara)

                    # kadar_kalori_per_kg (GCV)
                    list_gcv = list(FactSolverKontrakPemasok.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list('gcv_kcal_kg', flat=True))
                    print("\n Isi list GCV \n", list_gcv)

                    # FOB
                    list_fob = list(FactSolverKontrakPemasok.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list('fob_rp_mt', flat=True))

                    #  (CIF)
                    list_cif_cal = list(FactSolverKontrakPemasok.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list('cif_rp_cal', flat=True))

                    #  Sulphur
                    list_ts = list(FactSolverKontrakPemasok.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list('ts_persen', flat=True))

                    # Total kalori per KG
                    total_kalori_kg = sum(list_gcv)
                    print("\n Jumlah kalori per KG = ", total_kalori_kg)

                    list_jenis = list(FactSolverKontrakPemasok.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list('jenis_bb', flat=True))

                    # Nama Pemasok
                    list_pemasok_distinct = FactSolverKontrakPemasok.objects.filter(
                        pltu_id__nama_pltu__contains=nama_pltu).values_list('pemasok_id__nama_pemasok', flat=True).distinct()
                    print("\n Pemasok = ", list_pemasok_distinct)

                    # jenis per pemasok
                    list_jenis_distinct = []
                    for pemasok in list_pemasok_distinct:
                        print(pemasok)
                        jenis = FactSolverKontrakPemasok.objects.filter(
                            pltu_id__nama_pltu__contains=nama_pltu).filter(pemasok_id__nama_pemasok__contains=pemasok).values_list('pemasok_id__jenis_bb', flat=True).distinct().get()
                        print(jenis)
                        list_jenis_distinct.append(jenis)
                    print("\n isi list jenis distinct : \n", list_jenis_distinct)

                    # Total kalori per KG
                    total_kalori_kg = sum(list_gcv)
                    # print("\n Jumlah kalori per KG = ",total_kalori_kg)

                    list_pemasok_new = []
                    for pemasok in list_pemasok_distinct:
                        new_pemasok = pemasok.replace(" ", "_")
                        list_pemasok_new.append(new_pemasok)

                    print("\n LIST PEMASOK NEW :  \n", list_pemasok_new)

                    # membuat list volume resume per pemasok sepeti list_resume_volume_per_pemasok=[[Muara Enim,Muara Enim],[Kalsel,Kalsel,Kalsel],...]
                    count = 0
                    list_resume_volume_per_pemasok = []
                    list_resume_ts_per_pemasok = []
                    list_resume_harga_per_pemasok = []
                    list_resume_kalori_per_pemasok = []

                    for i, x in enumerate(list_pemasok_new):
                        globals()[x] = list(FactSolverKontrakPemasok.objects.filter(
                            pltu_id__nama_pltu__contains=nama_pltu).filter(pemasok_id__nama_pemasok=x.replace("_", " ")).values_list(
                            'nama_sumber_tambang', flat=True))
                        # print("\n INI GLOBAL before : \n", globals()[x])
                        list_resume_volume_per_pemasok.append(globals()[x])
                        list_resume_ts_per_pemasok.append(globals()[x])
                        list_resume_harga_per_pemasok.append(globals()[x])
                        list_resume_kalori_per_pemasok.append(globals()[x])

                    print('\n LIST RESUME VOLUME PER PEMASOK : \n',
                        list_resume_volume_per_pemasok)

                    print('\n LIST RESUME TS PER PEMASOKKK : \n',
                        list_resume_ts_per_pemasok)

                    print("\n panjang volume : \n = ", len(list_resume_volume_per_pemasok))
                    print("\n panjang ts : \n = ", len(list_resume_ts_per_pemasok))
                    print("\n panjang harga : \n = ", len(list_resume_harga_per_pemasok))

                    no = 0
                    list_resume_volume_per_pemasok_new = []

                    for i, vol in enumerate(list_resume_volume_per_pemasok):
                        # print(val)
                        for j in range(len(vol)):
                            # print(j)
                            vol[j] = list_volume[no]
                            no += 1
                        vol = lpSum(vol)

                        list_resume_volume_per_pemasok_new.append(vol)

                        # print("\n INI GLOBAL X : \n", globals()[x])
                    print('\n LIST RESUME VOLUME PER PEMASOK NEW: \n',
                        list_resume_volume_per_pemasok_new)

                    jumlah_resume_volume = lpSum(list_resume_volume_per_pemasok_new)

                    # list_resume_volume_per_pemasok = [bukit_asam, arutmin_reg, arutmin_csr, eei, hanson_lrc, hanson_mrc,
                    #                                 rap, titan_lrc, titan_mrc, kii, dizamatra, plnbb_lrc, plnbb_mrc, oktasan]

                    print("\n panjang pemasok resume: \n = ", len(list_pemasok_distinct))
                    print("\n panjang jenis : \n = ", len(list_jenis_distinct))
                    print("\n panjang volume : \n = ", len(list_resume_volume_per_pemasok_new))

                    df_resume_volume = pd.DataFrame(
                        {'Pemasok': list_pemasok_distinct,
                        'Jenis': list(list_jenis_distinct),
                        'Volume (MT)': list_resume_volume_per_pemasok_new,

                        })

                    print("\n DF Resume Solver python: \n = ", df_resume_volume)

                    LRC = df_resume_volume[df_resume_volume['Jenis'].str.contains(
                        'LRC')]['Volume (MT)']
                    MRC = df_resume_volume[df_resume_volume['Jenis'].str.contains(
                        'MRC')]['Volume (MT)']

                    print("\n LRC : \n  ", LRC)
                    print("\n MRC : \n  ", MRC)

                    total_LRC_resume = sum(LRC)
                    total_MRC_resume = sum(MRC)

                    print("\n total LRC : \n  ", total_LRC_resume)
                    print("\n total MRC : \n ", total_MRC_resume)

                    ''' +++++++++++++++++++++++++++++++++++++++++++ MODEL PROBLEM +++++++++++++++++++++++++++++++++++++++++ '''
                    # definisikan model minimum
                    prob = p.LpProblem("Minimum Biaya Batubara", p.LpMinimize)

                    # definisikan solusi/objective function

                    # GOALS - Meminimalkan biaya batubara
                    # mencari nilai R63 = Total Harga (Rp)
                    # Multiplying two lists

                    string_goal = ""
                    for i in range(0, len(list_cif_cal)):
                        string_goal += (list_cif_cal[i] * list_volume[i])

                    print("\n String goal = ", string_goal)
                    print("\n\n")
                    prob += string_goal

                    '''  ========================= CONSTRAIN ============================== '''
                    # Batasan

                    # BATASAN Persentase LRC dan MRC
                    # rumus awal "persentase_LRC=(total_LRC_resume/jumlah_resume_volume)*100"
                    # pulp tidak support pembagian, maka dijadikan perkalian =
					prob += persentase_batas_lrc*input_rencana_batubara >= total_LRC_resume*100
                    #prob += persentase_batas_lrc*jumlah_resume_volume <= total_LRC_resume*100
                    #prob += persentase_batas_mrc*jumlah_resume_volume <= total_MRC_resume*100


                    # BATASAN NILAI MINIMUM PASOKAN
                    for i, val in enumerate(list_resume_volume_per_pemasok_new):
                        prob += val >= 0

                    # BATASAN NILAI MAKSIMUM PASOKAN
                    for i, val in enumerate(list_resume_volume_per_pemasok_new):
                        prob += val <= list_max_batubara[i]

                    # batasan total_volume dan nilai kalor tertimbang
                    # input_nilai_kebutuhan_batubara =2981414
                    # input_nilai_kebutuhan_kalor =  4350

                    total_kalori_per_ton = ""
                    for i in range(0, len(list_volume)):
                        total_kalori_per_ton += list_volume[i] * list_gcv[i] * 1000

                    total_sulphur = ""
                    for i in range(0, len(list_volume)):
                        total_sulphur += list_volume[i] * list_ts[i]

                    prob += total_volume >= input_rencana_batubara

                    # nilai kalor tertimbang
                    # rumus awal = total_kalor_per_ton/(total_volume*1000)>= input_nilai_kebutuhan_kalor
                    # pulp tidak support pembagian
                    prob += total_kalori_per_ton >= input_rencana_kalor*(total_volume*1000)

                    # total sulphur %
                    # rumus awal = total_sulphur/(total_volume)>= input_rencana_kebutuhan_sulphur
                    #  tidak support pembagian karena masih dalam bentuk variabel x, belum ada nilai x nya
                    #ON OFF sulfur
                    if(statussulfur == "Pakai"):
                        prob += total_sulphur <= input_rencana_sulphur*(total_volume)

                    # prob.writeLP(
                    #    "C:\D\PROJECT_LANJUT\IP\django_solver\solver_batubara\solver\Model.lp")
                    # RUN Solver
                    status = prob.solve()
                    status_model = p.LpStatus[status]
                    print("\n\n\n ++++++++++++++++++++++++++++++++++++ |||||||||||||||| +++++++++++++++++++++++++++++++++")
                    print("\n ++++++++++++++++++++++++++++++++++++ RUN SOLVE MODEL +++++++++++++++++++++++++++++++++")
                    print("\n ++++++++++++++++++++++++++++++++++++ ||||||||||||||| +++++++++++++++++++++++++++++++++\n")
                    print("\n ---|||| STATUS model LP yang terbentuk = ",
                        status_model)
                    print("\n Nilai Objective model LP  = ", p.value(prob.objective))

                    status_solver = p.LpStatus[status]

                    list_hasil_volume = []
                    for a in list_volume:
                        list_hasil_volume.append(value(a))

                    print("\n list_hasil volume \n", list_hasil_volume)

                    # LIST HARGA PER SUMBER TAMBANG
                    list_harga = []
                    for cif, vol in zip(list_cif_cal, list_hasil_volume):
                        hasil_kali = cif*vol
                        list_harga.append(hasil_kali)

                    print("\n panjang list harga : \n", len(list_harga))
                    total_harga = sum(list_harga)

                    print("\n Total harga \n", total_harga)

                    print("\n ============================ ++++++++++++++++++++++++++ $$$$$$$$$$$$$$$$$$$$ ======================================== \n")

                    # membuat list volume per pemasok hasil dari setelah solver dijalankan.
                    no = 0
                    list_hasil_volume_per_pemasok = []
                    list_resume_ts_per_pemasok_new = []
                    list_resume_harga_per_pemasok_new = []
                    list_resume_kalori_per_pemasok_new = []

                    for i, val in enumerate(list_resume_volume_per_pemasok):

                        for j in range(len(val)):
                            # print(j)
                            val[j] = value(list_volume[no])
                    #     print("\n Volume : ", val)
                            no += 1

                    # print("\n  Volume bawah:", val)
                        val = lpSum(val)
                    # print("\n Jumlah volume :", val)

                        list_hasil_volume_per_pemasok.append(val)

                    print("\n $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$  \n")
                    no = 0
                    for i, ts in enumerate(list_resume_ts_per_pemasok):
                        for j in range(len(ts)):
                            ts[j] = list_hasil_volume[no]*list_ts[no]
                        #    print("\n TS:", ts)
                            no += 1

                        #print("\n  TS bawah:", ts)
                        jumlah_ts = sum(ts)
                        #print("\n Jumlah TS :", jumlah_ts)

                        list_resume_ts_per_pemasok_new.append(jumlah_ts)

                    print("\n $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$  \n")
                    no = 0
                    for i, kalori in enumerate(list_resume_kalori_per_pemasok):
                        for j in range(len(kalori)):
                            kalori[j] = list_gcv[no]*list_hasil_volume[no]*1000
                            no += 1

                        #print("\n  Kalori bawah:", kalori)
                        jumlah_kalori = sum(kalori)
                        #print("\n Jumlah kalori :", jumlah_kalori)

                        list_resume_kalori_per_pemasok_new.append(jumlah_kalori)

                    print("\n $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$  \n")

                    no = 0
                    for i, harga in enumerate(list_resume_harga_per_pemasok):
                        for j in range(len(harga)):
                            # print(j)
                            harga[j] = list_hasil_volume[no]*list_cif_cal[no]
                        #    print("\n Harga:", harga)
                            no += 1

                    # print("\n  harga bawah:", harga)
                        jumlah_harga = sum(harga)
                    # print("\n Jumlah Harga :", jumlah_harga)

                        list_resume_harga_per_pemasok_new.append(jumlah_harga)

                        print("\n ================================================================= \n")

                    print('\n LIST HASIL SOLVER per pemasok: \n',
                        list_hasil_volume_per_pemasok)

                    print('\n LIST HASIL TS per pemasok: \n',
                        list_resume_ts_per_pemasok_new)

                    print('\n LIST HASIL Harga per pemasok: \n',
                        list_resume_harga_per_pemasok_new)

                    print("\n panjang volume : \n = ", len(list_resume_volume_per_pemasok))
                    print("\n panjang ts : \n = ", len(list_resume_ts_per_pemasok_new))
                    print("\n panjang harga : \n = ", len(list_resume_harga_per_pemasok_new))

                    jumlah_hasil_resume_volume = lpSum(list_hasil_volume_per_pemasok)

                    hasil_total_volume_x = sum(list_hasil_volume_per_pemasok)

                    #  kalori X
                    list_kalori_x = []
                    for i in range(len(list_hasil_volume)):
                        kalori_x = list_gcv[i]*list_hasil_volume[i]*1000
                        list_kalori_x.append(kalori_x)

                    # list nilai sulphur hasil solver
                    list_ts_x = []
                    for i in range(len(list_hasil_volume)):
                        sulphur_x = list_ts[i]*list_hasil_volume[i]
                        list_ts_x.append(sulphur_x)

                    hasil_total_kalori_x = sum(list_kalori_x)

                    # total jumlah suplhur X
                    hasil_total_sulphur_per_pemasok = sum(list_ts_x)

                    print("\n hasil total sulphur per pemasok \n",
                        hasil_total_sulphur_per_pemasok)

                    print("\n list ts x \n", list_ts_x)

                    # list_jenis_per_pemasok=['MRC','LRC','LRC','LRC','LRC','MRC','LRC','LRC','MRC','LRC','MRC','LRC','MRC','LRC']
                    list_gabungan_resume_max = zip(list_pemasok_distinct, list_jenis_distinct,
                                                list_pemasok_distinct, list_max_batubara)

                    for i, val in enumerate(list_hasil_volume_per_pemasok):
                        print('-> besar Volume MT', list_pemasok_distinct[i], '  ', val)

                    df_resume_volume = pd.DataFrame(
                        {'Pemasok': list_pemasok_distinct,
                            'Jenis': list_jenis_distinct,
                            'Volume (MT)': [value(item) for item in list_hasil_volume_per_pemasok],
                            'TS': list_resume_ts_per_pemasok_new,
                            'Harga': list_resume_harga_per_pemasok_new,
                        })

                    print("\n DF Resume Solver python: \n = ", df_resume_volume)

                    print("\n List hasil volume per pemasok: \n = ",
                        list_hasil_volume_per_pemasok)

                    LRC = df_resume_volume[df_resume_volume['Jenis'].str.contains(
                        'LRC')]['Volume (MT)']
                    MRC = df_resume_volume[df_resume_volume['Jenis'].str.contains(
                        'MRC')]['Volume (MT)']

                    print("\n LRC : \n  ", LRC)
                    print("\n MRC : \n  ", MRC)

                    total_LRC_resume = sum(LRC)
                    total_MRC_resume = sum(MRC)

                    print("\n total LRC : \n  ", total_LRC_resume)
                    print("\n total MRC : \n ", total_MRC_resume)

                    print("\n Total Volume  = \n ", jumlah_hasil_resume_volume)

                    persentase_LRC = total_LRC_resume*100 / \
                        float(value(jumlah_hasil_resume_volume))
                    persentase_MRC = total_MRC_resume*100 / \
                        float(value(jumlah_hasil_resume_volume))

                    print("\n Persentase akhir LRC : \n  ", persentase_LRC, "%")
                    print("\n Persentase akhir MRC : \n ", persentase_MRC, "%")

                    jumlah_hasil_resume_volume = float(value(jumlah_hasil_resume_volume))

                    total_sulphur = hasil_total_sulphur_per_pemasok / \
                        jumlah_hasil_resume_volume

                    print("\n TOTAL SULPHUR (%) : \n  ", total_sulphur, " %")

                    # nilai kalor tertimbang
                    # sudah bisa di bagi, karena nilai x sudah didapatkan
                    nilai_kalor_tertimbang = hasil_total_kalori_x / \
                        (float(value(hasil_total_volume_x))*1000)

                    print("\n NILAI KALOR TERTIMBANG : \n  ", nilai_kalor_tertimbang)

                    rp_per_cal = float(value(total_harga)) / \
                        float(value(jumlah_hasil_resume_volume))

                    rp_per_mt = rp_per_cal/float(value(jumlah_hasil_resume_volume)) * \
                        (hasil_total_kalori_x/1000)

                    # ========================================================================
                    # =================================SIMPAN DATABASE========================
                    # ========================================================================

                    # per sumber tambang -> simpan ke tabel fact_solver_hasil_optimalisasisolver
                    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')

                    print("\n NOW \n", now)

                    FactSolverHasilSatuan.objects.create(
                        persentase_lrc_persen=persentase_LRC, persentase_mrc_persen=persentase_MRC,
                        volume_lrc_mt=total_LRC_resume, volume_mrc_mt=total_MRC_resume, total_volume_akhir=jumlah_hasil_resume_volume, rp_cal=rp_per_cal,
                        rp_mt=rp_per_mt, total_kalori_akhir=hasil_total_kalori_x, total_ts_akhir=hasil_total_sulphur_per_pemasok, nilai_kalor_kcal_kg=nilai_kalor_tertimbang,
                        total_sulphur_persen=float(value(total_sulphur)), status_model=status_model,
                        total_harga_akhir=float(value(total_harga)))

                    nomor_proses = FactSolverHasilSatuan.objects.last().noproses_id
                    print(nomor_proses)
                    '''
                    if(no_proses == 0):
                        nomor = 1
                    else:
                        nomor = no_proses+1
                    '''
                    # per sumber tambang -> simpan ke tabel fact_solver_hasil_optimalisasi

                    for i, (pemasok, sumbertambang, vol, ts, kal, harga) in enumerate(zip(list_id_pemasok,
                                                                                        list_sumber_tambang, list_hasil_volume, list_ts_x, list_kalori_x, list_harga)):
                        FactSolverHasilOptimalisasisolver.objects.create(
                            pltu_id=id_pltu, pemasok_id=pemasok, noproses_id=int(nomor_proses), nama_sumber_tambang=sumbertambang, waktu=now,
                            hasilsolver_volume_mt=vol,  hasilsolver_ts_persen=ts,  hasilsolver_kalori_kcal=kal,  hasilsolver_harga_rp_mt_cal=float(value(harga)))

                    # per pemasok -> simpan ke tabel fact_solver_resume_hasil_optimalisasi

                    '''
                    for i, vol in enumerate(list_hasil_volume_per_pemasok):
                        FactSolverResumeHasilOptimasi(
                            rhos_volume=float(value(vol))).save()


                    '''
                    for i, (id_pemasok, jenis, vol, kalori, ts, harga) in enumerate(
                        zip(list_id_pemasok_distinct, list_jenis_distinct, list_hasil_volume_per_pemasok,
                            list_resume_kalori_per_pemasok_new, list_resume_ts_per_pemasok_new, list_resume_harga_per_pemasok_new)):
                        FactSolverResumeHasilOptimasi(pltu_id=id_pltu, pemasok_id=id_pemasok, noproses_id=nomor_proses, waktu=now, jenis_bb=jenis, rhos_volume=float(value(vol)),
                                                    rhos_kalori_kcal=kalori, rhos_ts_persen=ts, rhos_harga_rp_mt_cal=harga).save()

                    #return render(request,'about/hasiloptimalisasi.html', context4)
                    return redirect('hasiloptimalisasi')
                else:
                    return render(request,'about/inputsolver.html', context4)
                #data = FactSolverRencana(pltu_id=a, presentase_rencana_mrc_persen=b, presentase_rencana_lrc_persen=c, input_bb_mt=d, input_nilai_kalor_kcal_kg=e, input_total_sulphur=f)
                #data.save()
                return render(request,'about/inputsolver.html', context4)
        else:
            #queryTableMasterSolver = FactSolverKontrakPemasok.objects.all().order_by('-id_kontrakpemasok')
            #queryTableMasterSolver = FactSolverKontrakPemasok.objects.all()
            queryPLTU = PLTU.objects.all()
            #queryPemasok = Pemasok.objects.all()
            #querySumberTambang = Sumbertambang.objects.all()
            #queryWaktu = Waktu.objects.all()
            #queryWaktuDistinctTriwulan = Waktu.objects.all().values('triwulan').distinct()
            #queryWaktuDistinctTahun = Waktu.objects.all().values('tahun').distinct()
            #mastersolverform = MasterSolverForm()
            #contact_form = ContactForm()
            context = {
                'judul':'Solver',
                'kontributor':'Admin',
                'nav': [
                    ['/','Home'],
                    ['/blog/news','News'],
                    ['/about','About'],
                ],
                'app_css':'blog/css/styles.css',
                #'mastersolverform':mastersolverform,
                #'contact_form':contact_form,
                #'dataMasterSolver':queryTableMasterSolver,
                'queryPLTU':queryPLTU,
                #'queryPemasok':queryPemasok,
                #'querySumberTambang':querySumberTambang,
                #'queryWaktu':queryWaktu,
                #'queryWaktuDistinctTriwulan':queryWaktuDistinctTriwulan,
                #'queryWaktuDistinctTahun':queryWaktuDistinctTahun,
            }

            return render(request,'about/inputsolver.html', context)
    else:
        return redirect('login')


def recent(request):
    return HttpResponse('<h1>ini recent</h1>')

def tambahmastersolver(request):
    if request.user.is_authenticated:
        if request.method == "POST":
            request.POST._mutable = True

            pltu = request.POST['pltu']
            query_get_id_pltu = PLTU.objects.get(nama_pltu=pltu)
            request.POST['pltu'] = query_get_id_pltu.pltu_id

            pemasok = request.POST['pemasok']
            query_get_id_pemasok = Pemasok.objects.get(nama_pemasok=pemasok)
            request.POST['pemasok'] = query_get_id_pemasok.pemasok_id

            #sumbertambang = request.POST['nama_sumber_tambang']
            #query_get_id_sumbertambang = Sumbertambang.objects.get(nama_sumber_tambang=sumbertambang)
            #request.POST['sumbertambang'] = query_get_id_sumbertambang.sumbertambang_id

            triwulan = request.POST['triwulan']
            tahun = request.POST['waktu']
            query_get_id_waktu = Waktu.objects.get(triwulan=triwulan, tahun=tahun)
            request.POST['waktu'] = query_get_id_waktu.waktu_id


            request.POST._mutable = False

            form = ValidasiMasterSolver(request.POST)
            if form.is_valid():
                form.save()
                return HttpResponseRedirect('/about')
            return HttpResponseRedirect('/about')
    else:
        return redirect('login')
def hapusmastersolver(request, pk):
    if request.user.is_authenticated:
        if request.method == "POST":
            if request.POST["logout"] == "":
                logout(request)
                return redirect('login')

        mastersolver = FactSolverKontrakPemasok.objects.get(pk=pk)
        mastersolver.delete()
        return HttpResponseRedirect('/about')
    else:
        return redirect('login')

def updatemastersolver(request, pk):
    if request.user.is_authenticated:
        '''if request.method == "POST":
            if request.POST["logout"] == "":
                logout(request)
                return redirect('login')'''
        if 'logout' in request.POST:
            logoutnya = request.POST['logout']
            if logoutnya == "":
                logout(request)
                return redirect('login')
        request.POST._mutable = True

        pltu = request.POST['pltu']
        query_get_id_pltu = PLTU.objects.get(nama_pltu=pltu)
        request.POST['pltu'] = query_get_id_pltu.pltu_id

        pemasok = request.POST['pemasok']
        query_get_id_pemasok = Pemasok.objects.get(nama_pemasok=pemasok)
        request.POST['pemasok'] = query_get_id_pemasok.pemasok_id

        triwulan = request.POST['triwulan']
        tahun = request.POST['waktu']
        query_get_id_waktu = Waktu.objects.get(triwulan=triwulan, tahun=tahun)
        request.POST['waktu'] = query_get_id_waktu.waktu_id

        request.POST._mutable = False

        if request.method == 'POST':
            instance = get_object_or_404(FactSolverKontrakPemasok, pk=pk)
            form = ValidasiMasterSolver(request.POST or None, instance=instance)
            if form.is_valid():
                form.save()
            return HttpResponseRedirect('/about')
        else:
            return HttpResponseRedirect('/about')
    else:
        return redirect('login')


def editmastersolver(request, pk):
    if request.user.is_authenticated:
        '''if request.method == "POST":
            if request.POST["logout"] == "":
                logout(request)
                return redirect('login')'''
        if 'logout' in request.POST:
            logoutnya = request.POST['logout']
            if logoutnya == "":
                logout(request)
                return redirect('login')

        #queryTableMasterSolver = FactSolverKontrakPemasok.objects.all()
        queryPLTU = PLTU.objects.all()
        queryPemasok = Pemasok.objects.all()
        querySumberTambang = Sumbertambang.objects.all()
        #queryWaktu = Waktu.objects.all()
        queryWaktuDistinctTriwulan = Waktu.objects.all().values('triwulan').distinct()
        queryWaktuDistinctTahun = Waktu.objects.all().values('tahun').distinct()
        queryJenisbbDistinct = Pemasok.objects.all().values('jenis_bb').distinct()

        querytwumum = Waktu.objects.filter(~Q(tahun=2019)).values('triwulan').distinct()
        querytw2019 = Waktu.objects.filter(Q(tahun=2019)).values('triwulan').distinct()
        querytahunumum = Waktu.objects.filter(~Q(tahun=2019)).values('tahun').distinct()


        #pltu = request.POST['id_pltu']
        #query_get_id_pltu = PLTU.objects.get(nama_pltu=pltu)
        #id_pltunya = query_get_id_pltu.pltu_id
        #queryPemasokDistinct = FactSolverKontrakPemasok.objects.filter(id_pltu=id_pltunya).values('id_pemasok').disinct()
        queryEditMasterSolver = FactSolverKontrakPemasok.objects.filter(pk=pk)
        context = {
            'judul':'Edit Solver',
            'kontributor':'Admin',
            'app_css':'blog/css/styles.css',
            #'dataMasterSolver':queryTableMasterSolver,
            'queryPLTU':queryPLTU,
            'queryPemasok':queryPemasok,
            'querySumberTambang':querySumberTambang,
            #'queryWaktu':queryWaktu,
            'queryWaktuDistinctTriwulan':queryWaktuDistinctTriwulan,
            'queryWaktuDistinctTahun':queryWaktuDistinctTahun,
            #'queryPemasokDistinct':queryPemasokDistinct,
            'queryEditMasterSolver':queryEditMasterSolver,
            'queryJenisbbDistinct':queryJenisbbDistinct,
            'querytwumum':querytwumum,
            'querytw2019':querytw2019,
            'querytahunumum':querytahunumum

        }

        return render(request,'about/editmastersolver.html', context)
    else:
        return redirect('login')

def hasiloptimalisasisolver(request):
    print(request.method)
    #print(request.GET)
    print(request.POST)
    if request.user.is_authenticated:
        '''if request.method == "POST":
            if request.POST["logout"] == "":
                logout(request)
                return redirect('login')'''
        if 'logout' in request.POST:
            logoutnya = request.POST['logout']
            if logoutnya == "":
                logout(request)
                return redirect('login')
        if 'sort' in request.POST:
            # GA KEPAKE NI FUNGSI IF INI, LONCAT KE ELSE
            sort = request.POST['sort']
            if sort != None:
                #queryTableMasterSolver = FactSolverKontrakPemasok.objects.all().order_by('-id_kontrakpemasok')
                queryTableMasterSolver = FactSolverKontrakPemasok.objects.all()
                queryPLTU = PLTU.objects.all()
                queryPemasok = Pemasok.objects.all()
                querySumberTambang = Sumbertambang.objects.all()
                queryWaktu = Waktu.objects.all()
                queryWaktuDistinctTriwulan = Waktu.objects.all().values('triwulan').distinct()
                queryWaktuDistinctTahun = Waktu.objects.all().values('tahun').distinct()
                querySumberTambangDistinct = Sumbertambang.objects.all().values('nama_sumber_tambang').distinct()

                sortpltu = request.POST['sortpltu']
                query_get_id_pltu = PLTU.objects.get(nama_pltu=sortpltu)
                id_pltunya = query_get_id_pltu.pltu_id
                querypltufilter = FactSolverKontrakPemasok.objects.filter(pltu = id_pltunya)
                #mastersolverform = MasterSolverForm()
                #contact_form = ContactForm()
                context = {
                    'judul':'Solver',
                    'kontributor':'Admin',
                    'nav': [
                        ['/','Home'],
                        ['/blog/news','News'],
                        ['/about','About'],
                    ],
                    'app_css':'blog/css/styles.css',
                    #'mastersolverform':mastersolverform,
                    #'contact_form':contact_form,
                    'dataMasterSolver':queryTableMasterSolver,
                    'queryPLTU':queryPLTU,
                    'queryPemasok':queryPemasok,
                    'querySumberTambang':querySumberTambang,
                    'queryWaktu':queryWaktu,
                    'queryWaktuDistinctTriwulan':queryWaktuDistinctTriwulan,
                    'queryWaktuDistinctTahun':queryWaktuDistinctTahun,
                    'querypltufilter':querypltufilter,
                    'querySumberTambangDistinct':querySumberTambangDistinct,

                }
                return render(request,'about/index.html', context)

        else:
            #queryTableMasterSolver = FactSolverKontrakPemasok.objects.all().order_by('-id_kontrakpemasok')
            #queryTableMasterSolver = FactSolverKontrakPemasok.objects.all()
            #queryPLTU = PLTU.objects.all()
            #queryPemasok = Pemasok.objects.all()
            #querySumberTambang = Sumbertambang.objects.all()
            #queryWaktu = Waktu.objects.all()
            #queryWaktuDistinctTriwulan = Waktu.objects.all().values('triwulan').distinct()
            #queryWaktuDistinctTahun = Waktu.objects.all().values('tahun').distinct()
            #querySumberTambangDistinct = Sumbertambang.objects.all().values('nama_sumber_tambang').distinct()

            checknorun = FactSolverHasilOptimalisasisolver.objects.order_by('-hasil_id')[0]
            queryHasil = FactSolverHasilOptimalisasisolver.objects.filter(noproses = checknorun.noproses)
            no = checknorun.noproses_id
            querySatuan = FactSolverHasilSatuan.objects.get(noproses_id = no)

            context = {
                'judul':'Solver',
                'kontributor':'Admin',
                'nav': [
                    ['/','Home'],
                    ['/blog/news','News'],
                    ['/about','About'],
                ],
                'app_css':'blog/css/styles.css',
                #'mastersolverform':mastersolverform,
                #'contact_form':contact_form,
                #'dataMasterSolver':queryTableMasterSolver,
                #'queryPLTU':queryPLTU,
                #'queryPemasok':queryPemasok,
                #'querySumberTambang':querySumberTambang,
                #'queryWaktu':queryWaktu,
                #'queryWaktuDistinctTriwulan':queryWaktuDistinctTriwulan,
                #'queryWaktuDistinctTahun':queryWaktuDistinctTahun,
                #'querySumberTambangDistinct':querySumberTambangDistinct,
                'queryHasil':queryHasil,
                'checknorun':checknorun,
                'querySatuan':querySatuan


            }

            return render(request,'about/hasiloptimalisasi.html', context)
    else:
        #return render(request, 'login.html')
        return redirect('login')

def resumesolver(request):
    print(request.method)
    #print(request.GET)
    print(request.POST)
    if request.user.is_authenticated:
        '''if request.method == "POST":
            if request.POST["logout"] == "":
                logout(request)
                return redirect('login')'''
        if 'logout' in request.POST:
            logoutnya = request.POST['logout']
            if logoutnya == "":
                logout(request)
                return redirect('login')
        if 'cetakexcel' in request.POST:

            cetakexcel = request.POST['cetakexcel']
            if cetakexcel != None:

                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
                response['Content-Disposition'] = 'attachment; filename={date}-resume_hasil_optimalisasi.xlsx'.format(date=datetime.datetime.now().strftime('%Y-%m-%d'),)
                workbook = Workbook()

                # Get active worksheet/tab
                worksheet = workbook.active
                worksheet.title = 'Resume Hasil Optimalisasi'

                checknorunHasil = FactSolverResumeHasilOptimasi.objects.order_by('-rhos_id')[0]
                no = checknorunHasil.noproses_id
                querySatuan = FactSolverHasilSatuan.objects.get(noproses_id = no)

                pltunya = checknorunHasil.pltu.nama_pltu
                waktunya = checknorunHasil.waktu

                cellpltu = worksheet.cell(row=1, column=1)
                cellpltu.value = pltunya
                cellwaktu = worksheet.cell(row=1, column=2)
                cellwaktu.value = waktunya


                # Define the titles for columns
                columns = [
                    'Pemasok', 'Jenis', 'Volume (MT)', 'Kalori (kCal)', 'Sulphur (%)', 'Harga (Rp*MT/Cal)'
                ]
                row_num = 2


                # Assign the titles for each cell of the header
                for col_num, column_title in enumerate(columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                queryHasil = FactSolverResumeHasilOptimasi.objects.filter(noproses = checknorunHasil.noproses)
                # Iterate through all rekaps
                for rekap in queryHasil:
                    row_num += 1

                    # Define the data for each cell in the row
                    row = [
                    rekap.pemasok.nama_pemasok,
                    rekap.jenis_bb,
                    rekap.rhos_volume,
                    rekap.rhos_kalori_kcal,
                    rekap.rhos_ts_persen,
                    rekap.rhos_harga_rp_mt_cal,

                    ]

                    # Assign the data for each cell of the row
                    for col_num, cell_value in enumerate(row, 1):
                        #dttm = datetime.strptime(row[14], "%m/%d/%Y")
                        cell = worksheet.cell(row=row_num, column=col_num)#.value = dttm
                        cell.value = cell_value



                rownumtotal = row_num + 1
                worksheet.merge_cells(start_row=rownumtotal, start_column=1, end_row=rownumtotal, end_column=2)
                cellmergetotaljadi = worksheet.cell(row=rownumtotal, column=1)
                cellmergetotaljadi.value = "Total: "
                cellmergetotaljadi.alignment = Alignment(horizontal='center', vertical='center')

                celltva = worksheet.cell(row=rownumtotal, column=3)
                celltva.value = querySatuan.total_volume_akhir

                celltka = worksheet.cell(row=rownumtotal, column=4)
                celltka.value = querySatuan.total_kalori_akhir

                celltta = worksheet.cell(row=rownumtotal, column=5)
                celltta.value = querySatuan.total_ts_akhir

                celltha = worksheet.cell(row=rownumtotal, column=6)
                celltha.value = querySatuan.total_harga_akhir

                rownumtotal2 = rownumtotal + 1
                worksheet.merge_cells(start_row=rownumtotal2, start_column=1, end_row=rownumtotal2, end_column=2)
                cellmergehargarpmt = worksheet.cell(row=rownumtotal2, column=1)
                cellmergehargarpmt.value = "Harga (Rp/MT): "
                cellmergehargarpmt.alignment = Alignment(horizontal='center', vertical='center')

                #volakhirsatuan = querySatuan.total_volume_akhir
                #rp_per_cal = querySatuan.rp_mt_setahun
                #totalcalakhir = querySatuan.total_kalori_akhir
                #rp_per_mt = rp_per_cal/volakhirsatuan*(totalcalakhir/1000)

                cellrpmt = worksheet.cell(row=rownumtotal2, column=3)
                cellrpmt.value = querySatuan.rp_mt

                rownumtotal3 = rownumtotal + 2
                worksheet.merge_cells(start_row=rownumtotal3, start_column=1, end_row=rownumtotal3, end_column=2)
                cellmergehargarpcal = worksheet.cell(row=rownumtotal3,column=1)
                cellmergehargarpcal.value = "Harga (Rp/Cal): "
                cellmergehargarpcal.alignment = Alignment(horizontal='center', vertical='center')

                cellrpcal = worksheet.cell(row=rownumtotal3, column=3)
                cellrpcal.value = querySatuan.rp_cal

                rownumtotal4 = rownumtotal + 3
                worksheet.merge_cells(start_row=rownumtotal4, start_column=1, end_row=rownumtotal4, end_column=2)
                cellmergekalort = worksheet.cell(row=rownumtotal4,column=1)
                cellmergekalort.value = "Nilai Kalor Tertimbang (kCal/kg): "
                cellmergekalort.alignment = Alignment(horizontal='center', vertical='center')

                cellkalort = worksheet.cell(row=rownumtotal4, column=3)
                cellkalort.value = querySatuan.nilai_kalor_kcal_kg

                rownumtotal5 = rownumtotal + 4
                worksheet.merge_cells(start_row=rownumtotal5, start_column=1, end_row=rownumtotal5, end_column=2)
                cellmergets = worksheet.cell(row=rownumtotal5,column=1)
                cellmergets.value = "Total Sulfur (%): "
                cellmergets.alignment = Alignment(horizontal='center', vertical='center')

                cellts = worksheet.cell(row=rownumtotal5, column=3)
                cellts.value = querySatuan.total_sulphur_persen

                rownumbawah = rownumtotal5 + 2
                cellrkb = worksheet.cell(row=rownumbawah, column=1)
                cellrkb.value = "Rencana Komposisi Batubara"
                cellrkb.alignment = Alignment(horizontal='center', vertical='center')

                cellprs = worksheet.cell(row=rownumbawah, column=2)
                cellprs.value = "Persentase (%)"
                cellprs.alignment = Alignment(horizontal='center', vertical='center')

                cellvmt = worksheet.cell(row=rownumbawah, column=3)
                cellvmt.value = "Volume (MT)"
                cellvmt.alignment = Alignment(horizontal='center', vertical='center')

                rownumbawah2 = rownumbawah + 1
                cellmrc = worksheet.cell(row=rownumbawah2, column=1)
                cellmrc.value = "MRC"
                cellmrc.alignment = Alignment(horizontal='center', vertical='center')

                cellmrcp = worksheet.cell(row=rownumbawah2, column=2)
                cellmrcp.value = querySatuan.persentase_mrc_persen

                cellmrcv = worksheet.cell(row=rownumbawah2, column=3)
                cellmrcv.value = querySatuan.volume_mrc_mt

                rownumbawah3 = rownumbawah + 2
                celllrc = worksheet.cell(row=rownumbawah3, column=1)
                celllrc.value = "LRC"
                celllrc.alignment = Alignment(horizontal='center', vertical='center')

                celllrcp = worksheet.cell(row=rownumbawah3, column=2)
                celllrcp.value = querySatuan.persentase_lrc_persen

                celllrcv = worksheet.cell(row=rownumbawah3, column=3)
                celllrcv.value = querySatuan.volume_lrc_mt

                workbook.save(response)

                return response

        else:
            #queryTableMasterSolver = FactSolverKontrakPemasok.objects.all().order_by('-id_kontrakpemasok')
            #queryTableMasterSolver = FactSolverKontrakPemasok.objects.all()
            #queryPLTU = PLTU.objects.all()
            #queryPemasok = Pemasok.objects.all()
            #querySumberTambang = Sumbertambang.objects.all()
            #queryWaktu = Waktu.objects.all()
            #queryWaktuDistinctTriwulan = Waktu.objects.all().values('triwulan').distinct()
            #queryWaktuDistinctTahun = Waktu.objects.all().values('tahun').distinct()
            #querySumberTambangDistinct = Sumbertambang.objects.all().values('nama_sumber_tambang').distinct()

            checknorunHasil = FactSolverResumeHasilOptimasi.objects.order_by('-rhos_id')[0]
            queryHasil = FactSolverResumeHasilOptimasi.objects.filter(noproses = checknorunHasil.noproses)
            no = checknorunHasil.noproses_id
            querySatuan = FactSolverHasilSatuan.objects.get(noproses_id = no)

            #volakhirsatuan = querySatuan.total_volume_akhir
            #rp_per_cal = querySatuan.rp_mt_setahun
            #totalcalakhir = querySatuan.total_kalori_akhir
            #rp_per_mt = rp_per_cal/volakhirsatuan*(totalcalakhir/1000)

            context = {
                'judul':'Solver',
                'kontributor':'Admin',
                'nav': [
                    ['/','Home'],
                    ['/blog/news','News'],
                    ['/about','About'],
                ],
                'app_css':'blog/css/styles.css',
                #'mastersolverform':mastersolverform,
                #'contact_form':contact_form,
                #'dataMasterSolver':queryTableMasterSolver,
                #'queryPLTU':queryPLTU,
                #'queryPemasok':queryPemasok,
                #'querySumberTambang':querySumberTambang,
                #'queryWaktu':queryWaktu,
                #'queryWaktuDistinctTriwulan':queryWaktuDistinctTriwulan,
                #'queryWaktuDistinctTahun':queryWaktuDistinctTahun,
                #'querySumberTambangDistinct':querySumberTambangDistinct,
                'queryHasil':queryHasil,
                'checknorunHasil':checknorunHasil,
                'querySatuan':querySatuan,
                #'rp_per_mt':rp_per_mt

            }

            return render(request,'about/resumesolver.html', context)
    else:
        #return render(request, 'login.html')
        return redirect('login')
