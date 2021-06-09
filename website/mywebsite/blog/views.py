from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from master.models import *
from .models import *
from .forms import *
from datetime import datetime, timedelta
from django.db.models import Sum
import csv
from openpyxl import Workbook
from django.contrib import messages

# Create your views here.
def loginView(request):
    print("Masuk1",request.method)
    #print(request.GET)
    print("Masuk1a",request.POST)
    context = {
        'page_title':'LOGIN',
    }
    user = None

    if request.method == "GET":
        if request.user.is_authenticated:
            # logika untuk user
            print("Masuk2", request.method)
            #print(request.GET)
            print("masuk2a",request.GET)
            return redirect('index')
        else:
            # logika untuk anonymous
            #return render(request, 'login.html', context)
            print("Masuk3", request.method)
            #print(request.GET)
            print("masuk3a",request.GET)
            return redirect('login')


    if request.method == "POST":
        print("Masuk4", request.method)
        #print(request.GET)
        print("masuk4a",request.POST)
        username_login = request.POST['username']
        password_login = request.POST['password']

        user = authenticate(request, username=username_login, password=password_login)

        if user is not None:
            print("Masuk5", request.method)
            #print(request.GET)
            print("masuk5a",request.POST)
            login(request, user)
            return redirect('index')
        else:
            print("Masuk6", request.method)
            #print(request.GET)
            print("masuk6a",request.POST)
            return redirect('login')

    return redirect('login')

# Create your views here.
def index(request):
    if request.user.is_authenticated:
        if 'logout' in request.POST:
            logoutnya = request.POST['logout']
            if logoutnya == "":
                logout(request)
                return redirect('login')

        if 'carirekap' in request.POST:
            carirekap = request.POST['carirekap']
            pltu = request.POST['pltu']
            tgl_awal = request.POST['tgl_awal']
            tgl_awal = datetime.strptime(tgl_awal, '%d/%m/%Y').strftime('%Y-%m-%d %H:%M')
            tgl_akhir = request.POST['tgl_akhir']
            tgl_akhir = datetime.strptime(tgl_akhir, '%d/%m/%Y').strftime('%Y-%m-%d 23:59')
            kolom=request.POST['kolom']
            print(kolom)
            if carirekap != None:
                if kolom == "time_arrival":
                    hasil1 = Rekap_shipment.objects.all().filter(pltu=pltu,time_arrival__gte=tgl_awal,time_arrival__lte=tgl_akhir).order_by('-no','-voyage')
                    print("berhasil 1")
                if kolom == "konfirmasi_rakor":
                    hasil1 = Rekap_shipment.objects.all().filter(pltu=pltu,konfirmasi_rakor__gte=tgl_awal,konfirmasi_rakor__lte=tgl_akhir).order_by('-no','-voyage')
                    print("berhasil 2")
                if kolom == "completed_unloading":
                    hasil1 = Rekap_shipment.objects.all().filter(pltu=pltu,completed_unloading__gte=tgl_awal,completed_unloading__lte=tgl_akhir).order_by('-no','-voyage')
                    print("berhasil 3")

                hasil2 = PLTU.objects.all()
                hasil3 = Pemasok.objects.all()
                hasil4 = Sumbertambang.objects.all()
                blog_form = BlogForm()
                contact_form = ContactForm()
                context = {
                    'judul':'Rekap Shipment Batu Bara',
                    'kontributor':'Admin',
                    'nav': [
                        ['/','Home'],
                        ['/blog/news','News'],
                        ['/about','About'],
                    ],
                    'app_css':'blog/css/styles.css',
                    'blog_form':blog_form,
                    'contact_form':contact_form,
                    'data1':hasil1,
                    'data2':hasil2,
                    'data3':hasil3,
                    'data4':hasil4
                }
                return render(request,'blog/index.html', context)

        if 'export' in request.POST:
            export = request.POST['export']
            pltu = request.POST['pltu']
            tgl_awal = request.POST['tgl_awal']
            tgl_awal = datetime.strptime(tgl_awal, '%d/%m/%Y').strftime('%Y-%m-%d %H:%M')
            tgl_akhir = request.POST['tgl_akhir']
            tgl_akhir = datetime.strptime(tgl_akhir, '%d/%m/%Y').strftime('%Y-%m-%d 23:59')
            kolom=request.POST['kolom']
            print(kolom)
            if export != None:
                if kolom == "time_arrival":
                    rekap_queryset = Rekap_shipment.objects.all().filter(pltu=pltu,time_arrival__gte=tgl_awal,time_arrival__lte=tgl_akhir).order_by('no','voyage')
                    print("berhasil 1")
                if kolom == "konfirmasi_rakor":
                    rekap_queryset = Rekap_shipment.objects.all().filter(pltu=pltu,konfirmasi_rakor__gte=tgl_awal,konfirmasi_rakor__lte=tgl_akhir).order_by('no','voyage')
                    print("berhasil 2")
                if kolom == "commenced_unloading":
                    rekap_queryset = Rekap_shipment.objects.all().filter(pltu=pltu,commenced_unloading__gte=tgl_awal,commenced_unloading__lte=tgl_akhir).order_by('no','voyage')
                    print("berhasil 3")

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename=Rekapshipment-{pltu}({date}).xlsx'.format(
                date=datetime.now().strftime('%Y-%m-%d'),
                pltu = request.POST['pltu']
            )
            workbook = Workbook()

            # Get active worksheet/tab
            worksheet = workbook.active
            worksheet.title = 'rekap Shipment'

            # Define the titles for columns
            columns = [
                'pltu', 'no', 'periode', 'bbo_id', 'shipment_code', 'voyage_code', 'voyage', 'nama_pemasok', 'transhipment_shifting', 'sumber_tambang', 'loading_port', 'unloading_port', 'tb', 'bg', 'konfirmasi_rakor', 'time_arrival', 'status_laycan', 'batas_maks_not_accepted', 'nor_accepted_pltu', 'despatch_demurrage_keterangan','despatch_demurrage_biaya', 'berthed_time', 'commenced_unloading', 'start_cleaning', 'completed_unloading','cast_off', 'gangguan_jam_coal_handling_facility', 'gangguan_jam_kendala_eksternal', 'gangguan_jam_force_majeur', 'gangguan_jam_gangguan_pengurang_flowrate', 'volume_b_l','volume_ds','initial_draught_surve_start','initial_draught_surve_finish', 'lift_off_loader_start', 'lift_off_loader_finish', 'final_draught_survey_start', 'final_draught_survey_fiinish','port_stay_day', 'port_stay_hour', 'berthing_time', 'initial_draught', 'waiting_commenced', 'persiapan_bongkar', 'lift_off_loader', 'final_draught', 'waiting_cast_off', 'persiapan_cast_off', 'waktu_tunggu_sandar', 'waktu_bongkar', 'waktu_cleaning', 'flowrate_gross_mt_hours', 'flowrate_nett_mt_hours', 'status_umpire', 'roaloading_gcv_arb_kcal_kg', 'roaloading_tm_arb', 'roaloading_im_adb', 'roaloading_ash_content_arb', 'roaloading_ash_content_adb', 'roaloading_vm_adb', 'roaloading_fc_adb', 'roaloading_total_sulphur_arb', 'roaloading_total_sulphur_dafb', 'roaloading_hgi_index', 'coaloading_gcv_arb_kcal_kg', 'coaloading_gcv_adb_kcal_kg', 'coaloading_gcv_db_kcal_kg', 'coaloading_gcv_dafb_kcal_kg', 'coaloading_tm_arb', 'coaloading_im_adb', 'coaloading_ash_content_arb', 'coaloading_ash_content_adb', 'coaloading_ash_content_db', 'coaloading_vm_arb', 'coaloading_vm_adb', 'coaloading_vm_db', 'coaloading_vm_dafb', 'coaloading_fc_arb', 'coaloading_fc_adb', 'coaloading_fc_db', 'coaloading_fc_dafb', 'coaloading_total_sulphur_arb', 'coaloading_total_sulphur_adb', 'coaloading_total_sulphur_db', 'coaloading_total_sulphur_dafb', 'coaloading_hgi_index', 'coaloading_idt_reducing_c', 'coaunloading_gcv_arb_kcal_kg', 'coaunloading_gcv_adb_kcal_kg', 'coaunloading_gcv_db_kcal_kg', 'coaunloading_gcv_dafb_kcal_kg', 'coaunloading_tm_arb', 'coaunloading_im_adb', 'coaunloading_ash_content_arb', 'coaunloading_ash_content_adb', 'coaunloading_ash_content_db', 'coaunloading_vm_arb', 'coaunloading_vm_adb', 'coaunloading_vm_db', 'coaunloading_vm_dafb', 'coaunloading_fc_arb', 'coaunloading_fc_adb', 'coaunloading_fc_db', 'coaunloading_fc_dafb', 'coaunloading_total_sulphur_arb', 'coaunloading_total_sulphur_adb', 'coaunloading_total_sulphur_db', 'coaunloading_total_sulphur_dafb', 'coaunloading_c_adb', 'coaunloading_h_adb', 'coaunloading_n_adb', 'coaunloading_n_dafb', 'coaunloading_o_adb', 'coaunloading_hgi_index', 'coaunloading_sizing_70mm', 'coaunloading_sizing_50mm', 'coaunloading_sizing_32mm', 'coaunloading_sizing_2_38mm', 'coaunloading_ashanalysis_sio2', 'coaunloading_ashanalysis_ai2o3', 'coaunloading_ashanalysis_fe2o3', 'coaunloading_ashanalysis_cao', 'coaunloading_ashanalysis_mgo', 'coaunloading_ashanalysis_na2o', 'coaunloading_ashanalysis_k2o', 'coaunloading_ashanalysis_tio2', 'coaunloading_ashanalysis_so3', 'coaunloading_ashanalysis_mno2', 'coaunloading_ashanalysis_p2o5', 'coaunloading_slagging_factor', 'coaunloading_fouling_factor', 'coaunloading_idt_reducing_c', 'coaunloading_ht_reducing_c', 'coaunloading_ht_oxidizing_c', 'coaunloading_hg_arb', 'labip_gcv_arb_kcal_kg', 'labip_gcv_adb_kcal_kg', 'labip_gcv_db_kcal_kg', 'labip_gcv_dafb', 'labip_tm_arb', 'labip_im_adb', 'labip_ash_content_arb', 'labip_ash_content_adb', 'labip_ash_content_db', 'labip_vm_arb', 'labip_vm_adb', 'labip_vm_db', 'labip_vm_dafb', 'labip_fc_arb', 'labip_fc_adb', 'labip_fc_db', 'labip_fc_dafb', 'labip_total_sulphur_arb', 'labip_total_sulphur_adb', 'labip_total_sulphur_db', 'labip_total_sulphur_dafb', 'labip_c_adb', 'labip_h_adb', 'labip_n_adb', 'labip_n_dafb', 'labip_o_adb', 'labip_hgi_index', 'quicktest_gcv_arb_kcal_kg', 'quicktest_tm_arb', 'quicktest_ahs_content_arb', 'quicktest_total_sulphur_arb'
            ]
            row_num = 1


            # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title

            #rekap_queryset = Rekap_shipment.objects.all()
            # Iterate through all rekaps
            for rekap in rekap_queryset:
                row_num += 1

                # Define the data for each cell in the row
                row = [
                rekap.pltu,
                rekap.no,
                rekap.periode,
                rekap.bbo_id,
                rekap.shipment_code,
                rekap.voyage_code,
                rekap.voyage,
                rekap.nama_pemasok,
                rekap.transhipment_shifting,
                rekap.sumber_tambang,
                rekap.loading_port,
                rekap.unloading_port,
                rekap.tb,
                rekap.bg,
                rekap.konfirmasi_rakor,
                rekap.time_arrival,
                rekap.status_laycan,
                rekap.batas_maks_not_accepted,
                rekap.nor_accepted_pltu,
                rekap.despatch_demurrage_keterangan,
                rekap.despatch_demurrage_biaya,
                rekap.berthed_time,
                rekap.commenced_unloading,
                rekap.start_cleaning,
                rekap.completed_unloading,
                rekap.cast_off,
                rekap.gangguan_jam_coal_handling_facility,
                rekap.gangguan_jam_kendala_eksternal,
                rekap.gangguan_jam_force_majeur,
                rekap.gangguan_jam_gangguan_pengurang_flowrate,
                rekap.volume_b_l,
                rekap.volume_ds,
                rekap.initial_draught_surve_start,
                rekap.initial_draught_surve_finish,
                rekap.lift_off_loader_start,
                rekap.lift_off_loader_finish,
                rekap.final_draught_survey_start,
                rekap.final_draught_survey_fiinish,
                rekap.port_stay_day,
                rekap.port_stay_hour,
                rekap.berthing_time,
                rekap.initial_draught,
                rekap.waiting_commenced,
                rekap.persiapan_bongkar,
                rekap.lift_off_loader,
                rekap.final_draught,
                rekap.waiting_cast_off,
                rekap.persiapan_cast_off,
                rekap.waktu_tunggu_sandar,
                rekap.waktu_bongkar,
                rekap.waktu_cleaning,
                rekap.flowrate_gross_mt_hours,
                rekap.flowrate_nett_mt_hours,
                rekap.status_umpire,
                rekap.roaloading_gcv_arb_kcal_kg,
                rekap.roaloading_tm_arb,
                rekap.roaloading_im_adb,
                rekap.roaloading_ash_content_arb,
                rekap.roaloading_ash_content_adb,
                rekap.roaloading_vm_adb,
                rekap.roaloading_fc_adb,
                rekap.roaloading_total_sulphur_arb,
                rekap.roaloading_total_sulphur_dafb,
                rekap.roaloading_hgi_index,
                rekap.coaloading_gcv_arb_kcal_kg,
                rekap.coaloading_gcv_adb_kcal_kg,
                rekap.coaloading_gcv_db_kcal_kg,
                rekap.coaloading_gcv_dafb_kcal_kg,
                rekap.coaloading_tm_arb,
                rekap.coaloading_im_adb,
                rekap.coaloading_ash_content_arb,
                rekap.coaloading_ash_content_adb,
                rekap.coaloading_ash_content_db,
                rekap.coaloading_vm_arb,
                rekap.coaloading_vm_adb,
                rekap.coaloading_vm_db,
                rekap.coaloading_vm_dafb,
                rekap.coaloading_fc_arb,
                rekap.coaloading_fc_adb,
                rekap.coaloading_fc_db,
                rekap.coaloading_fc_dafb,
                rekap.coaloading_total_sulphur_arb,
                rekap.coaloading_total_sulphur_adb,
                rekap.coaloading_total_sulphur_db,
                rekap.coaloading_total_sulphur_dafb,
                rekap.coaloading_hgi_index,
                rekap.coaloading_idt_reducing_c,
                rekap.coaunloading_gcv_arb_kcal_kg,
                rekap.coaunloading_gcv_adb_kcal_kg,
                rekap.coaunloading_gcv_db_kcal_kg,
                rekap.coaunloading_gcv_dafb_kcal_kg,
                rekap.coaunloading_tm_arb,
                rekap.coaunloading_im_adb,
                rekap.coaunloading_ash_content_arb,
                rekap.coaunloading_ash_content_adb,
                rekap.coaunloading_ash_content_db,
                rekap.coaunloading_vm_arb,
                rekap.coaunloading_vm_adb,
                rekap.coaunloading_vm_db,
                rekap.coaunloading_vm_dafb,
                rekap.coaunloading_fc_arb,
                rekap.coaunloading_fc_adb,
                rekap.coaunloading_fc_db,
                rekap.coaunloading_fc_dafb,
                rekap.coaunloading_total_sulphur_arb,
                rekap.coaunloading_total_sulphur_adb,
                rekap.coaunloading_total_sulphur_db,
                rekap.coaunloading_total_sulphur_dafb,
                rekap.coaunloading_c_adb,
                rekap.coaunloading_h_adb,
                rekap.coaunloading_n_adb,
                rekap.coaunloading_n_dafb,
                rekap.coaunloading_o_adb,
                rekap.coaunloading_hgi_index,
                rekap.coaunloading_sizing_70mm,
                rekap.coaunloading_sizing_50mm,
                rekap.coaunloading_sizing_32mm,
                rekap.coaunloading_sizing_2_38mm,
                rekap.coaunloading_ashanalysis_sio2,
                rekap.coaunloading_ashanalysis_ai2o3,
                rekap.coaunloading_ashanalysis_fe2o3,
                rekap.coaunloading_ashanalysis_cao,
                rekap.coaunloading_ashanalysis_mgo,
                rekap.coaunloading_ashanalysis_na2o,
                rekap.coaunloading_ashanalysis_k2o,
                rekap.coaunloading_ashanalysis_tio2,
                rekap.coaunloading_ashanalysis_so3,
                rekap.coaunloading_ashanalysis_mno2,
                rekap.coaunloading_ashanalysis_p2o5,
                rekap.coaunloading_slagging_factor,
                rekap.coaunloading_fouling_factor,
                rekap.coaunloading_idt_reducing_c,
                rekap.coaunloading_ht_reducing_c,
                rekap.coaunloading_ht_oxidizing_c,
                rekap.coaunloading_hg_arb,
                rekap.labip_gcv_arb_kcal_kg,
                rekap.labip_gcv_adb_kcal_kg,
                rekap.labip_gcv_db_kcal_kg,
                rekap.labip_gcv_dafb,
                rekap.labip_tm_arb,
                rekap.labip_im_adb,
                rekap.labip_ash_content_arb,
                rekap.labip_ash_content_adb,
                rekap.labip_ash_content_db,
                rekap.labip_vm_arb,
                rekap.labip_vm_adb,
                rekap.labip_vm_db,
                rekap.labip_vm_dafb,
                rekap.labip_fc_arb,
                rekap.labip_fc_adb,
                rekap.labip_fc_db,
                rekap.labip_fc_dafb,
                rekap.labip_total_sulphur_arb,
                rekap.labip_total_sulphur_adb,
                rekap.labip_total_sulphur_db,
                rekap.labip_total_sulphur_dafb,
                rekap.labip_c_adb,
                rekap.labip_h_adb,
                rekap.labip_n_adb,
                rekap.labip_n_dafb,
                rekap.labip_o_adb,
                rekap.labip_hgi_index,
                rekap.quicktest_gcv_arb_kcal_kg,
                rekap.quicktest_tm_arb,
                rekap.quicktest_ahs_content_arb,
                rekap.quicktest_total_sulphur_arb,
                ]

                # Assign the data for each cell of the row
                for col_num, cell_value in enumerate(row, 1):
                    #dttm = datetime.strptime(row[14], "%m/%d/%Y")
                    cell = worksheet.cell(row=row_num, column=col_num)#.value = dttm
                    cell.value = cell_value

            workbook.save(response)

            return response


        else:
            hasil2 = PLTU.objects.all()
            hasil3 = Pemasok.objects.all()
            hasil4 = Sumbertambang.objects.all()
            blog_form = BlogForm()
            contact_form = ContactForm()
            context = {
                'judul':'Rekap Shipment Batu Bara',
                'kontributor':'Admin',
                'app_css':'blog/css/styles.css',
                'blog_form':blog_form,
                'contact_form':contact_form,
                'data2':hasil2,
                'data3':hasil3,
                'data4':hasil4
            }
            return render(request,'blog/index.html', context)
    else:
        #return render(request, 'login.html')
        return redirect('login')

def recent(request):
    return HttpResponse('<h1>ini recent</h1>')

def tambahrekap(request):

    if request.user.is_authenticated:
        if 'logout' in request.POST:
            logoutnya = request.POST['logout']
            if logoutnya == "":
                logout(request)
                return redirect('login')

        request.POST._mutable = True

        #rumus periode
        strperiode = request.POST['nor_accepted_pltu']
        request.POST['periode'] = datetime.strptime(strperiode, '%d/%m/%Y %H:%M').strftime('%b-%y')

        #shipment_code = request.POST['shipment_code']
        pltu = request.POST['pltu']
        codeshipment = PLTU.objects.values('kode_pltu').filter(nama_pltu=pltu)
        for shipmentcode in codeshipment:
            print(shipmentcode['kode_pltu'])

        shipment_code = str(shipmentcode['kode_pltu'])

        #rumus shipment_code
        strtime_arrival = request.POST['time_arrival']
        tahun = datetime.strptime(strtime_arrival, '%d/%m/%Y %H:%M').strftime('%Y')
        bulan = datetime.strptime(strtime_arrival, '%d/%m/%Y %H:%M').strftime('%m')
        request.POST['shipment_code'] = shipment_code+" "+tahun+" "+bulan+" #"+request.POST['no']

        #rumus voyage
        pltu = request.POST['pltu']
        pemasok = request.POST['nama_pemasok']
        pemasokid = Pemasok.objects.values('pemasok_id').filter(nama_pemasok=pemasok)
        for idpemasok in pemasokid:
            print(idpemasok['pemasok_id'])
        request.POST['pemasok']=idpemasok['pemasok_id']
        jumlahvoyage = Rekap_shipment.objects.filter(pltu=pltu).filter(nama_pemasok=pemasok).count()
        request.POST['voyage'] = str(jumlahvoyage+1)

        #rumus voyage_code
        request.POST['voyage_code'] = shipment_code+" "+request.POST['nama_pemasok']+" #"+request.POST['voyage']

        #rumus status_laycan &  batas_maks_not_accepted
        laycan1 = request.POST['time_arrival']
        laycan2 = request.POST['konfirmasi_rakor']
        date_format = '%d/%m/%Y %H:%M'
        time_arrival = datetime.strptime(str(laycan1), date_format)
        konfirmasi_rakor = datetime.strptime(str(laycan2), date_format)
        laycan = time_arrival - konfirmasi_rakor
        if laycan.days>=-3 and laycan.days<=4:
            request.POST['status_laycan']="ON TIME"
            request.POST['batas_maks_not_accepted']=time_arrival
        elif laycan.days<-3:
            request.POST['status_laycan']="TIBA LEBIH CEPAT"
            request.POST['batas_maks_not_accepted']=konfirmasi_rakor
        else :
            request.POST['status_laycan']="Telat"
            request.POST['batas_maks_not_accepted'] = time_arrival + timedelta(days=7)

        #rumus despatch_demurrage_keterangan
        volume_ds =  float(request.POST['volume_ds'])
        cast_off = request.POST['cast_off']
        nor_accepted_pltu = request.POST['nor_accepted_pltu']
        date_format = '%d/%m/%Y %H:%M'
        cast_off_time = datetime.strptime(cast_off, date_format)
        nor_accepted_pltu_time = datetime.strptime(nor_accepted_pltu, date_format)
        time = datetime.strptime(str(cast_off_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(nor_accepted_pltu_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60

        biaya=(float(minutes)/1440)

        if volume_ds < 20000:
            if biaya>4:
                request.POST['despatch_demurrage_keterangan']="DEMURRAGE"
            else :
                request.POST['despatch_demurrage_keterangan']="DESPATCH"
        else :
            if biaya>6:
                request.POST['despatch_demurrage_keterangan']="DEMURRAGE"
            else :
                request.POST['despatch_demurrage_keterangan']="DESPATCH"


        #Rumus despatch_demurrage_biaya
        if request.POST['despatch_demurrage_keterangan']=="DESPATCH":
            if volume_ds < 20000:
                request.POST['despatch_demurrage_biaya'] = (4-biaya)*7500000
            else :
                request.POST['despatch_demurrage_biaya'] = (6-biaya)*7500000

        elif request.POST['despatch_demurrage_keterangan']=="DEMURRAGE":
            if volume_ds < 20000:
                request.POST['despatch_demurrage_biaya'] =  (biaya-4)*-15000000
            else :
                request.POST['despatch_demurrage_biaya'] = (biaya-6)*-15000000

        #rumus gangguan_jam_gangguan_pengurang_flowrate
        request.POST['gangguan_jam_gangguan_pengurang_flowrate'] = float(request.POST['gangguan_jam_kendala_eksternal']) + float(request.POST['gangguan_jam_force_majeur'])

        #rumus port_stay_day
        cast_off = request.POST['cast_off']
        time_arrival = request.POST['time_arrival']
        date_format = '%d/%m/%Y %H:%M'
        cast_off_time = datetime.strptime(cast_off, date_format)
        time_arrival_time = datetime.strptime(time_arrival, date_format)
        time = datetime.strptime(str(cast_off_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(time_arrival_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['port_stay_day']=(float(minutes)/1440)
        request.POST['port_stay_hour']=request.POST['port_stay_day']*24

        #rumus berthing_time
        cast_off = request.POST['cast_off']
        berthed_time = request.POST['berthed_time']
        date_format = '%d/%m/%Y %H:%M'
        cast_off_time = datetime.strptime(cast_off, date_format)
        berthed_time_time = datetime.strptime(berthed_time, date_format)
        time = datetime.strptime(str(cast_off_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(berthed_time_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['berthing_time']=round((float(minutes)/1440)*24,2)

        #rumus initial_draught

        initial_draught_surve_start = request.POST['initial_draught_surve_start']
        initial_draught_surve_finish = request.POST['initial_draught_surve_finish']
        date_format = '%d/%m/%Y %H:%M'
        initial_draught_surve_start_time = datetime.strptime(initial_draught_surve_start, date_format)
        initial_draught_surve_finish_time = datetime.strptime(initial_draught_surve_finish, date_format)
        time = datetime.strptime(str(initial_draught_surve_finish_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(initial_draught_surve_start_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['initial_draught']=round((float(minutes)/1440)*24,2)

        #rumus waiting_commenced
        commenced_unloading = request.POST['commenced_unloading']
        initial_draught_surve_finish = request.POST['initial_draught_surve_finish']
        date_format = '%d/%m/%Y %H:%M'
        commenced_unloading_time = datetime.strptime(commenced_unloading, date_format)
        initial_draught_surve_finish_time = datetime.strptime(initial_draught_surve_finish, date_format)
        time = datetime.strptime(str(commenced_unloading_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(initial_draught_surve_finish_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['waiting_commenced']=round((float(minutes)/1440)*24,2)

        #rumus persiapan_bongkar
        request.POST['persiapan_bongkar']=request.POST['initial_draught']+request.POST['waiting_commenced']

        #rumus lift_off_loader
        lift_off_loader_finish = request.POST['lift_off_loader_finish']
        lift_off_loader_start = request.POST['lift_off_loader_start']
        print(lift_off_loader_finish+" "+lift_off_loader_start+" hasil2")
        if lift_off_loader_finish=='' or lift_off_loader_start=='':
            print("kosong")
            request.POST['lift_off_loader']=0
        else:
            date_format = '%d/%m/%Y %H:%M'
            lift_off_loader_start_time = datetime.strptime(lift_off_loader_start, date_format)
            lift_off_loader_finish_time = datetime.strptime(lift_off_loader_finish, date_format)
            time = datetime.strptime(str(lift_off_loader_finish_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(lift_off_loader_start_time), '%Y-%m-%d %H:%M:%S')
            minutes = time.total_seconds() / 60
            request.POST['lift_off_loader']=round((float(minutes)/1440)*24,2)

        #rumus final_draught
        final_draught_survey_fiinish = request.POST['final_draught_survey_fiinish']
        final_draught_survey_start = request.POST['final_draught_survey_start']
        date_format = '%d/%m/%Y %H:%M'
        final_draught_survey_start_time = datetime.strptime(final_draught_survey_start, date_format)
        final_draught_survey_fiinish_time = datetime.strptime(final_draught_survey_fiinish, date_format)
        time = datetime.strptime(str(final_draught_survey_fiinish_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(final_draught_survey_start_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['final_draught']=round((float(minutes)/1440)*24,2)



        #rumus waiting_cast_off
        cast_off = request.POST['cast_off']
        final_draught_survey_fiinish = request.POST['final_draught_survey_fiinish']
        date_format = '%d/%m/%Y %H:%M'
        cast_off_time = datetime.strptime(cast_off, date_format)
        final_draught_survey_fiinish_time = datetime.strptime(final_draught_survey_fiinish, date_format)
        time = datetime.strptime(str(cast_off_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(final_draught_survey_fiinish_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['waiting_cast_off']=round((float(minutes)/1440)*24,2)

        #rumus persiapan_cast_off
        request.POST['persiapan_cast_off'] = request.POST['lift_off_loader'] + request.POST['final_draught'] + request.POST['waiting_cast_off']

        #rumus waktu_tunggu_sandar
        berthed_time = request.POST['berthed_time']
        nor_accepted_pltu = request.POST['nor_accepted_pltu']
        date_format = '%d/%m/%Y %H:%M'
        berthed_time_time = datetime.strptime(berthed_time, date_format)
        nor_accepted_pltu_time = datetime.strptime(nor_accepted_pltu, date_format)
        time = datetime.strptime(str(berthed_time_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(nor_accepted_pltu_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['waktu_tunggu_sandar']=round((float(minutes)/1440)*24,2)

        #rumus waktu_bongkar
        commenced_unloading = request.POST['commenced_unloading']
        completed_unloading = request.POST['completed_unloading']
        date_format = '%d/%m/%Y %H:%M'
        commenced_unloading_time = datetime.strptime(commenced_unloading, date_format)
        completed_unloading_time = datetime.strptime(completed_unloading, date_format)
        time = datetime.strptime(str(completed_unloading_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(commenced_unloading_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['waktu_bongkar']=round((float(minutes)/1440)*24,2)

        #rumus waktu_cleaning
        start_cleaning = request.POST['start_cleaning']
        completed_unloading = request.POST['completed_unloading']
        date_format = '%d/%m/%Y %H:%M'
        start_cleaning_time = datetime.strptime(start_cleaning, date_format)
        completed_unloading_time = datetime.strptime(completed_unloading, date_format)
        time = datetime.strptime(str(completed_unloading_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(start_cleaning_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['waktu_cleaning']=round((float(minutes)/1440)*24,2)

        #rumus flowrate_gross_mt_hours
        volume_ds=float(request.POST['volume_ds'])
        bagi = float(request.POST['persiapan_bongkar']+request.POST['persiapan_cast_off']+request.POST['waktu_bongkar'])
        request.POST['flowrate_gross_mt_hours']=round(float(volume_ds/bagi),2)

        #rumus flowrate_nett_mt_hours
        volume_ds=float(request.POST['volume_ds'])
        bagi = float(request.POST['waktu_bongkar'])-float(request.POST['gangguan_jam_kendala_eksternal'])-float(request.POST['gangguan_jam_force_majeur'])
        request.POST['flowrate_nett_mt_hours']=round(float(volume_ds/bagi),2)

        #rumus coaloading_gcv_arb_kcal_kg
        coaloading_gcv_adb_kcal_kg=float(request.POST['coaloading_gcv_adb_kcal_kg'])
        coaloading_tm_arb=float(request.POST['coaloading_tm_arb'])
        coaloading_im_adb=float(request.POST['coaloading_im_adb'])
        request.POST['coaloading_gcv_arb_kcal_kg']=round(coaloading_gcv_adb_kcal_kg*(100-coaloading_tm_arb)/(100-coaloading_im_adb), 2)

        #rumus coaloading_gcv_db_kcal_kg
        coaloading_gcv_adb_kcal_kg=float(request.POST['coaloading_gcv_adb_kcal_kg'])
        coaloading_im_adb=float(request.POST['coaloading_im_adb'])
        request.POST['coaloading_gcv_db_kcal_kg']=round((coaloading_gcv_adb_kcal_kg*100)/(100-coaloading_im_adb), 2)

        #rumus coaloading_gcv_dafb_kcal_kg
        coaloading_ash_content_adb=float((request.POST['coaloading_ash_content_adb']))
        request.POST['coaloading_gcv_dafb_kcal_kg']=round(int(coaloading_gcv_adb_kcal_kg*100/(100-coaloading_im_adb-coaloading_ash_content_adb)), 2)

        #rumus coaloading_ash_content_arb
        request.POST['coaloading_ash_content_arb']=round(coaloading_ash_content_adb*(100-coaloading_tm_arb)/(100-coaloading_im_adb), 2)

        #rumus coaloading_ash_content_db
        request.POST['coaloading_ash_content_db']=round((coaloading_ash_content_adb*100)/(100-coaloading_im_adb), 2)

        #rumus coaloading_vm_arb
        coaloading_vm_adb=float(request.POST['coaloading_vm_adb'])
        request.POST['coaloading_vm_arb']=round(coaloading_vm_adb*(100-coaloading_tm_arb)/(100-coaloading_im_adb), 2)

        #rumus coaloading_vm_db
        request.POST['coaloading_vm_db']=round((coaloading_vm_adb*100)/(100-coaloading_im_adb), 2)

        #rumus coaloading_vm_dafb
        coaloading_ash_content_adb=float((request.POST['coaloading_ash_content_adb']))
        request.POST['coaloading_vm_dafb']=round(float(coaloading_vm_adb*100/(100-coaloading_im_adb-coaloading_ash_content_adb)), 2)

        #rumus coaloading_fc_arb
        coaloading_fc_adb=float(request.POST['coaloading_fc_adb'])
        request.POST['coaloading_fc_arb']=round(coaloading_fc_adb*(100-coaloading_tm_arb)/(100-coaloading_im_adb), 2)

        #rumus coaloading_fc_db
        request.POST['coaloading_fc_db']=round((coaloading_fc_adb*100)/(100-coaloading_im_adb), 2)

        #rumus coaloading_fc_dafb
        request.POST['coaloading_fc_dafb']=round((coaloading_fc_adb*100/(100-coaloading_im_adb-coaloading_ash_content_adb)), 2)

        #rumus coaloading_total_sulphur_arb
        coaloading_total_sulphur_adb=float(request.POST['coaloading_total_sulphur_adb'])
        request.POST['coaloading_total_sulphur_arb']=round(coaloading_total_sulphur_adb*(100-coaloading_tm_arb)/(100-coaloading_im_adb), 2)

        #rumus coaloading_total_sulphur_db
        request.POST['coaloading_total_sulphur_db']=round((coaloading_total_sulphur_adb*100)/(100-coaloading_im_adb), 2)

        #rumus coaloading_total_sulphur_dafb
        request.POST['coaloading_total_sulphur_dafb']=round((coaloading_total_sulphur_adb*100/(100-coaloading_im_adb-coaloading_ash_content_adb)), 2)

        #rumus coaunloading_gcv_arb_kcal_kg
        coaunloading_gcv_adb_kcal_kg=float(request.POST['coaunloading_gcv_adb_kcal_kg'])
        coaunloading_tm_arb=float(request.POST['coaunloading_tm_arb'])
        coaunloading_im_adb=float(request.POST['coaunloading_im_adb'])
        request.POST['coaunloading_gcv_arb_kcal_kg']=round(coaunloading_gcv_adb_kcal_kg*(100-coaunloading_tm_arb)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_gcv_db_kcal_kg
        coaunloading_gcv_adb_kcal_kg=float(request.POST['coaunloading_gcv_adb_kcal_kg'])
        coaunloading_im_adb=float(request.POST['coaunloading_im_adb'])
        request.POST['coaunloading_gcv_db_kcal_kg']=round((coaunloading_gcv_adb_kcal_kg*100)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_gcv_dafb_kcal_kg
        coaunloading_ash_content_adb=float((request.POST['coaunloading_ash_content_adb']))
        request.POST['coaunloading_gcv_dafb_kcal_kg']=round(int(coaunloading_gcv_adb_kcal_kg*100/(100-coaunloading_im_adb-coaunloading_ash_content_adb)), 2)

        #rumus coaunloading_ash_content_arb
        request.POST['coaunloading_ash_content_arb']=round(coaunloading_ash_content_adb*(100-coaunloading_tm_arb)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_ash_content_db
        request.POST['coaunloading_ash_content_db']=round((coaunloading_ash_content_adb*100)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_vm_arb
        coaunloading_vm_adb=float(request.POST['coaunloading_vm_adb'])
        request.POST['coaunloading_vm_arb']=round(coaunloading_vm_adb*(100-coaunloading_tm_arb)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_vm_db
        request.POST['coaunloading_vm_db']=round((coaunloading_vm_adb*100)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_vm_dafb
        coaunloading_ash_content_adb=float((request.POST['coaunloading_ash_content_adb']))
        request.POST['coaunloading_vm_dafb']=round(float(coaunloading_vm_adb*100/(100-coaunloading_im_adb-coaunloading_ash_content_adb)), 2)

        #rumus coaunloading_fc_arb
        coaunloading_fc_adb=float(request.POST['coaunloading_fc_adb'])
        request.POST['coaunloading_fc_arb']=round(coaunloading_fc_adb*(100-coaunloading_tm_arb)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_fc_db
        request.POST['coaunloading_fc_db']=round((coaunloading_fc_adb*100)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_fc_dafb
        request.POST['coaunloading_fc_dafb']=round((coaunloading_fc_adb*100/(100-coaunloading_im_adb-coaunloading_ash_content_adb)), 2)

        #rumus coaunloading_total_sulphur_arb
        coaunloading_total_sulphur_adb=float(request.POST['coaunloading_total_sulphur_adb'])
        request.POST['coaunloading_total_sulphur_arb']=round(coaunloading_total_sulphur_adb*(100-coaunloading_tm_arb)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_total_sulphur_db
        request.POST['coaunloading_total_sulphur_db']=round((coaunloading_total_sulphur_adb*100)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_total_sulphur_dafb
        request.POST['coaunloading_total_sulphur_dafb']=round((coaunloading_total_sulphur_adb*100/(100-coaunloading_im_adb-coaunloading_ash_content_adb)), 2)

        #rumus coaunloading_n_dafb
        coaunloading_n_adb=float(request.POST['coaunloading_n_adb'])
        request.POST['coaunloading_n_dafb']=round((coaunloading_n_adb*100/(100-coaunloading_im_adb-coaunloading_ash_content_adb)), 2)

        #rumus coaunloading_o_adb
        coaunloading_c_adb=float(request.POST['coaunloading_c_adb'])
        coaunloading_h_adb=float(request.POST['coaunloading_h_adb'])
        request.POST['coaunloading_o_adb']=round((100-(coaunloading_im_adb+coaunloading_ash_content_adb+coaunloading_total_sulphur_adb+coaunloading_c_adb+coaunloading_h_adb+coaunloading_n_adb)), 2)

        #rumus labip_gcv_arb_kcal_kg
        labip_gcv_adb_kcal_kg=float(request.POST['labip_gcv_adb_kcal_kg'])
        labip_tm_arb=float(request.POST['labip_tm_arb'])
        labip_im_adb=float(request.POST['labip_im_adb'])
        request.POST['labip_gcv_arb_kcal_kg']=round(labip_gcv_adb_kcal_kg*(100-labip_tm_arb)/(100-labip_im_adb), 2)

        #rumus labip_gcv_db_kcal_kg
        labip_gcv_adb_kcal_kg=float(request.POST['labip_gcv_adb_kcal_kg'])
        labip_im_adb=float(request.POST['labip_im_adb'])
        request.POST['labip_gcv_db_kcal_kg']=round((labip_gcv_adb_kcal_kg*100)/(100-labip_im_adb), 2)

        #rumus labip_gcv_dafb
        labip_ash_content_adb=float((request.POST['labip_ash_content_adb']))
        request.POST['labip_gcv_dafb']=round(labip_gcv_adb_kcal_kg*100/(100-labip_im_adb-labip_ash_content_adb), 2)

        #rumus labip_ash_content_arb
        request.POST['labip_ash_content_arb']=round(labip_ash_content_adb*(100-labip_tm_arb)/(100-labip_im_adb), 2)

        #rumus labip_ash_content_db
        request.POST['labip_ash_content_db']=round((labip_ash_content_adb*100)/(100-labip_im_adb), 2)

        #rumus labip_vm_arb
        labip_vm_adb=float(request.POST['labip_vm_adb'])
        request.POST['labip_vm_arb']=round(labip_vm_adb*(100-labip_tm_arb)/(100-labip_im_adb), 2)

        #rumus labip_vm_db
        request.POST['labip_vm_db']=round((labip_vm_adb*100)/(100-labip_im_adb),2)

        #rumus labip_vm_dafb
        labip_ash_content_adb=float((request.POST['labip_ash_content_adb']))
        request.POST['labip_vm_dafb']=round(float(labip_vm_adb*100/(100-labip_im_adb-labip_ash_content_adb)), 2)

        #rumus labip_fc_arb
        labip_fc_adb=float(request.POST['labip_fc_adb'])
        request.POST['labip_fc_arb']=round(labip_fc_adb*(100-labip_tm_arb)/(100-labip_im_adb), 2)

        #rumus labip_fc_db
        request.POST['labip_fc_db']=round((labip_fc_adb*100)/(100-labip_im_adb), 2)

        #rumus labip_fc_dafb
        request.POST['labip_fc_dafb']=round((labip_fc_adb*100/(100-labip_im_adb-labip_ash_content_adb)), 2)

        #rumus labip_total_sulphur_arb
        labip_total_sulphur_adb=float(request.POST['labip_total_sulphur_adb'])
        request.POST['labip_total_sulphur_arb']=round(labip_total_sulphur_adb*(100-labip_tm_arb)/(100-labip_im_adb), 2)

        #rumus labip_total_sulphur_db
        request.POST['labip_total_sulphur_db']=round((labip_total_sulphur_adb*100)/(100-labip_im_adb), 2)

        #rumus labip_total_sulphur_dafb
        request.POST['labip_total_sulphur_dafb']=round((labip_total_sulphur_adb*100/(100-labip_im_adb-labip_ash_content_adb)), 2)

        #rumus labip_n_dafb
        labip_n_adb=float(request.POST['labip_n_adb'])
        request.POST['labip_n_dafb']=round((labip_n_adb*100/(100-labip_im_adb-labip_ash_content_adb)), 2)

        #rumus labip_o_adb
        labip_c_adb=float(request.POST['labip_c_adb'])
        labip_h_adb=float(request.POST['labip_h_adb'])
        #request.POST['labip_o_adb']=(100-(labip_im_adb+labip_ash_content_adb+labip_total_sulphur_adb+labip_c_adb+labip_h_adb+labip_n_adb))

        request.POST._mutable = False

        if request.method == 'POST':
            form = FormRekap(request.POST)

            if form.is_valid():
                form.save()
                print("berhasilllll")
                messages.info(request, 'Data Berhasil Disimpan!')
            else:
                print(form.errors)
            return HttpResponseRedirect('/blog')
        else:

            return HttpResponseRedirect('/blog')
    else:
        #return render(request, 'login.html')
        return redirect('login')

def hapusrekap(request, pk):
    rekap = Rekap_shipment.objects.get(pk=pk)
    rekap.delete()
    return HttpResponseRedirect('/blog')

def tambah(request):

    if request.user.is_authenticated:
        if 'logout' in request.POST:
            logoutnya = request.POST['logout']
            if logoutnya == "":
                logout(request)
                return redirect('login')

        hasil2 = PLTU.objects.all()
        hasil3 = Pemasok.objects.all()
        hasil4 = Sumbertambang.objects.all()
        context = {
            'judul':'Rekap Shipment Batu Bara',
            'kontributor':'Admin',
            'app_css':'blog/css/styles.css',
            'data2':hasil2,
            'data3':hasil3,
            'data4':hasil4,
        }
        return render(request,'blog/tambahrekap.html', context)
    else:
        #return render(request, 'login.html')
        return redirect('login')

def editrekap(request, pk):

    if request.user.is_authenticated:
        if request.method == "POST":
            if request.POST["logout"] == "":
                logout(request)
                return redirect('login')

        hasil1 = Rekap_shipment.objects.filter(pk=pk)
        hasil2 = PLTU.objects.all()
        hasil3 = Pemasok.objects.all()
        hasil4 = Sumbertambang.objects.all()
        hasil5 = Rekap_shipment.objects.filter(pk=pk)
        context = {
            'judul':'Rekap Shipment',
            'kontributor':'Admin',
            'app_css':'blog/css/styles.css',
            'data1':hasil1,
            'data2':hasil2,
            'data3':hasil3,
            'data4':hasil4,
            'data5':hasil5,
        }
        return render(request,'blog/editrekap.html', context)
    else:
        #return render(request, 'login.html')
        return redirect('login')

def updaterekap(request,pk):

    request.POST._mutable = True
    status_umpire=request.POST['status_umpire']

    if status_umpire=='UMPIRE':

        pemasok = request.POST['nama_pemasok']
        pemasokid = Pemasok.objects.values('pemasok_id').filter(nama_pemasok=pemasok)
        for idpemasok in pemasokid:
            print(idpemasok['pemasok_id'])
        request.POST['pemasok']=idpemasok['pemasok_id']
        request.POST._mutable = False

        if request.method == 'POST':
            #form = FormRekap(request.POST)

            instance = get_object_or_404(Rekap_shipment, pk=pk)
            form = FormRekap(request.POST or None, instance=instance)

            if form.is_valid():
                print(status_umpire)
                form.save()
            else:
                print(form.errors)
            return HttpResponseRedirect('/blog')
        else:
            return HttpResponseRedirect('/blog')

    else :

        #rumus periode
        strperiode = request.POST['nor_accepted_pltu']
        request.POST['periode'] = datetime.strptime(strperiode, '%d/%m/%Y %H:%M').strftime('%b-%y')
        #shipment_code = request.POST['shipment_code']

        #rumus shipment_code
        #strtime_arrival = request.POST['time_arrival']
        #tahun = datetime.strptime(strtime_arrival, '%d/%m/%Y %H:%M').strftime('%Y')
        #bulan = datetime.strptime(strtime_arrival, '%d/%m/%Y %H:%M').strftime('%m')
        #request.POST['shipment_code'] = request.POST['shipment_code']

        #rumus voyage
        #pltu = request.POST.get('pltu',True)
        #pemasok = request.POST['nama_pemasok']
        pemasok = request.POST['nama_pemasok']
        pemasokid = Pemasok.objects.values('pemasok_id').filter(nama_pemasok=pemasok)
        for idpemasok in pemasokid:
            print(idpemasok['pemasok_id'])
        request.POST['pemasok']=idpemasok['pemasok_id']
        #jumlahvoyage = Rekap_shipment.objects.filter(pltu=pltu).filter(nama_pemasok=pemasok).count()
        request.POST['voyage'] = request.POST['voyage']

        #rumus voyage_code
        request.POST['voyage_code'] = request.POST['voyage_code']

        #rumus status_laycan &  batas_maks_not_accepted
        laycan1 = request.POST['time_arrival']
        laycan2 = request.POST['konfirmasi_rakor']
        date_format = '%d/%m/%Y %H:%M'
        time_arrival = datetime.strptime(str(laycan1), date_format)
        konfirmasi_rakor = datetime.strptime(str(laycan2), date_format)
        laycan = time_arrival - konfirmasi_rakor
        if laycan.days>=-3 and laycan.days<=4:
            request.POST['status_laycan']="ON TIME"
            request.POST['batas_maks_not_accepted']=time_arrival
        elif laycan.days<-3:
            request.POST['status_laycan']="TIBA LEBIH CEPAT"
            request.POST['batas_maks_not_accepted']=konfirmasi_rakor
        else :
            request.POST['status_laycan']="Telat"
            request.POST['batas_maks_not_accepted'] = time_arrival + timedelta(days=7)

        #rumus despatch_demurrage_keterangan
        volume_ds =  float(request.POST['volume_ds'])
        cast_off = request.POST['cast_off']
        nor_accepted_pltu = request.POST['nor_accepted_pltu']
        date_format = '%d/%m/%Y %H:%M'
        cast_off_time = datetime.strptime(cast_off, date_format)
        nor_accepted_pltu_time = datetime.strptime(nor_accepted_pltu, date_format)
        time = datetime.strptime(str(cast_off_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(nor_accepted_pltu_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds()/60
        print("menit")

        biaya=(float(minutes)/1440)
        print(biaya)

        if volume_ds < 20000:
            if biaya>4:
                request.POST['despatch_demurrage_keterangan']="DEMURRAGE"
            else :
                request.POST['despatch_demurrage_keterangan']="DESPATCH"
        else :
            if biaya>6:
                request.POST['despatch_demurrage_keterangan']="DEMURRAGE"
            else :
                request.POST['despatch_demurrage_keterangan']="DESPATCH"


        #Rumus despatch_demurrage_biaya
        if request.POST['despatch_demurrage_keterangan']=="DESPATCH":
            if volume_ds < 20000:
                request.POST['despatch_demurrage_biaya'] = (4-biaya)*7500000
            else :
                request.POST['despatch_demurrage_biaya'] = (6-biaya)*7500000

        elif request.POST['despatch_demurrage_keterangan']=="DEMURRAGE":
            if volume_ds < 20000:
                request.POST['despatch_demurrage_biaya'] =  (biaya-4)*-15000000
            else :
                request.POST['despatch_demurrage_biaya'] = (biaya-6)*-15000000

        #rumus gangguan_jam_gangguan_pengurang_flowrate
        request.POST['gangguan_jam_gangguan_pengurang_flowrate'] = float(request.POST['gangguan_jam_kendala_eksternal']) + float(request.POST['gangguan_jam_force_majeur'])

        #rumus port_stay_day
        cast_off = request.POST['cast_off']
        time_arrival = request.POST['time_arrival']
        date_format = '%d/%m/%Y %H:%M'
        cast_off_time = datetime.strptime(cast_off, date_format)
        time_arrival_time = datetime.strptime(time_arrival, date_format)
        time = datetime.strptime(str(cast_off_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(time_arrival_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['port_stay_day']=(float(minutes)/1440)
        request.POST['port_stay_hour']=request.POST['port_stay_day']*24

        #rumus berthing_time
        cast_off = request.POST['cast_off']
        berthed_time = request.POST['berthed_time']
        date_format = '%d/%m/%Y %H:%M'
        cast_off_time = datetime.strptime(cast_off, date_format)
        berthed_time_time = datetime.strptime(berthed_time, date_format)
        time = datetime.strptime(str(cast_off_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(berthed_time_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['berthing_time']=round((float(minutes)/1440)*24,2)

        #rumus initial_draught

        initial_draught_surve_start = request.POST['initial_draught_surve_start']
        initial_draught_surve_finish = request.POST['initial_draught_surve_finish']
        date_format = '%d/%m/%Y %H:%M'
        initial_draught_surve_start_time = datetime.strptime(initial_draught_surve_start, date_format)
        initial_draught_surve_finish_time = datetime.strptime(initial_draught_surve_finish, date_format)
        time = datetime.strptime(str(initial_draught_surve_finish_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(initial_draught_surve_start_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['initial_draught']=round((float(minutes)/1440)*24,2)

        #rumus waiting_commenced
        commenced_unloading = request.POST['commenced_unloading']
        initial_draught_surve_finish = request.POST['initial_draught_surve_finish']
        date_format = '%d/%m/%Y %H:%M'
        commenced_unloading_time = datetime.strptime(commenced_unloading, date_format)
        initial_draught_surve_finish_time = datetime.strptime(initial_draught_surve_finish, date_format)
        time = datetime.strptime(str(commenced_unloading_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(initial_draught_surve_finish_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['waiting_commenced']=round((float(minutes)/1440)*24,2)

        #rumus persiapan_bongkar
        request.POST['persiapan_bongkar']=request.POST['initial_draught']+request.POST['waiting_commenced']

        #rumus lift_off_loader
        lift_off_loader_finish = request.POST['lift_off_loader_finish']
        lift_off_loader_start = request.POST['lift_off_loader_start']
        print(lift_off_loader_finish+" "+lift_off_loader_start+" hasil2")
        if lift_off_loader_finish=='' or lift_off_loader_start=='':
            print("kosong")
            request.POST['lift_off_loader']=0
        else:
            date_format = '%d/%m/%Y %H:%M'
            lift_off_loader_start_time = datetime.strptime(lift_off_loader_start, date_format)
            lift_off_loader_finish_time = datetime.strptime(lift_off_loader_finish, date_format)
            time = datetime.strptime(str(lift_off_loader_finish_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(lift_off_loader_start_time), '%Y-%m-%d %H:%M:%S')
            minutes = time.total_seconds() / 60
            request.POST['lift_off_loader']=round((float(minutes)/1440)*24,2)

        #rumus final_draught
        final_draught_survey_fiinish = request.POST['final_draught_survey_fiinish']
        final_draught_survey_start = request.POST['final_draught_survey_start']
        date_format = '%d/%m/%Y %H:%M'
        final_draught_survey_start_time = datetime.strptime(final_draught_survey_start, date_format)
        final_draught_survey_fiinish_time = datetime.strptime(final_draught_survey_fiinish, date_format)
        time = datetime.strptime(str(final_draught_survey_fiinish_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(final_draught_survey_start_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['final_draught']=round((float(minutes)/1440)*24,2)



        #rumus waiting_cast_off
        cast_off = request.POST['cast_off']
        final_draught_survey_fiinish = request.POST['final_draught_survey_fiinish']
        date_format = '%d/%m/%Y %H:%M'
        cast_off_time = datetime.strptime(cast_off, date_format)
        final_draught_survey_fiinish_time = datetime.strptime(final_draught_survey_fiinish, date_format)
        time = datetime.strptime(str(cast_off_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(final_draught_survey_fiinish_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['waiting_cast_off']=round((float(minutes)/1440)*24,2)

        #rumus persiapan_cast_off
        request.POST['persiapan_cast_off'] = float(request.POST['lift_off_loader']) + request.POST['final_draught'] + request.POST['waiting_cast_off']

        #rumus waktu_tunggu_sandar
        berthed_time = request.POST['berthed_time']
        nor_accepted_pltu = request.POST['nor_accepted_pltu']
        date_format = '%d/%m/%Y %H:%M'
        berthed_time_time = datetime.strptime(berthed_time, date_format)
        nor_accepted_pltu_time = datetime.strptime(nor_accepted_pltu, date_format)
        time = datetime.strptime(str(berthed_time_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(nor_accepted_pltu_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['waktu_tunggu_sandar']=round((float(minutes)/1440)*24,2)

        #rumus waktu_bongkar
        commenced_unloading = request.POST['commenced_unloading']
        completed_unloading = request.POST['completed_unloading']
        date_format = '%d/%m/%Y %H:%M'
        commenced_unloading_time = datetime.strptime(commenced_unloading, date_format)
        completed_unloading_time = datetime.strptime(completed_unloading, date_format)
        time = datetime.strptime(str(completed_unloading_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(commenced_unloading_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['waktu_bongkar']=round((float(minutes)/1440)*24,2)

        #rumus waktu_cleaning
        start_cleaning = request.POST['start_cleaning']
        completed_unloading = request.POST['completed_unloading']
        date_format = '%d/%m/%Y %H:%M'
        start_cleaning_time = datetime.strptime(start_cleaning, date_format)
        completed_unloading_time = datetime.strptime(completed_unloading, date_format)
        time = datetime.strptime(str(completed_unloading_time), '%Y-%m-%d %H:%M:%S') - datetime.strptime(str(start_cleaning_time), '%Y-%m-%d %H:%M:%S')
        minutes = time.total_seconds() / 60
        request.POST['waktu_cleaning']=round((float(minutes)/1440)*24,2)

        #rumus flowrate_gross_mt_hours
        volume_ds=float(request.POST['volume_ds'])
        bagi = float(request.POST['persiapan_bongkar']+request.POST['persiapan_cast_off']+request.POST['waktu_bongkar'])
        request.POST['flowrate_gross_mt_hours']=round(float(volume_ds/bagi),2)

        #rumus flowrate_nett_mt_hours
        volume_ds=float(request.POST['volume_ds'])
        bagi = float(request.POST['waktu_bongkar'])-float(request.POST['gangguan_jam_kendala_eksternal'])-float(request.POST['gangguan_jam_force_majeur'])
        request.POST['flowrate_nett_mt_hours']=round(float(volume_ds/bagi),2)

        #rumus coaloading_gcv_arb_kcal_kg
        coaloading_gcv_adb_kcal_kg=float(request.POST['coaloading_gcv_adb_kcal_kg'])
        coaloading_tm_arb=float(request.POST['coaloading_tm_arb'])
        coaloading_im_adb=float(request.POST['coaloading_im_adb'])
        request.POST['coaloading_gcv_arb_kcal_kg']=round(coaloading_gcv_adb_kcal_kg*(100-coaloading_tm_arb)/(100-coaloading_im_adb), 2)

        #rumus coaloading_gcv_db_kcal_kg
        coaloading_gcv_adb_kcal_kg=float(request.POST['coaloading_gcv_adb_kcal_kg'])
        coaloading_im_adb=float(request.POST['coaloading_im_adb'])
        request.POST['coaloading_gcv_db_kcal_kg']=round((coaloading_gcv_adb_kcal_kg*100)/(100-coaloading_im_adb), 2)

        #rumus coaloading_gcv_dafb_kcal_kg
        coaloading_ash_content_adb=float((request.POST['coaloading_ash_content_adb']))
        request.POST['coaloading_gcv_dafb_kcal_kg']=round(int(coaloading_gcv_adb_kcal_kg*100/(100-coaloading_im_adb-coaloading_ash_content_adb)), 2)

        #rumus coaloading_ash_content_arb
        request.POST['coaloading_ash_content_arb']=round(coaloading_ash_content_adb*(100-coaloading_tm_arb)/(100-coaloading_im_adb), 2)

        #rumus coaloading_ash_content_db
        request.POST['coaloading_ash_content_db']=round((coaloading_ash_content_adb*100)/(100-coaloading_im_adb), 2)

        #rumus coaloading_vm_arb
        coaloading_vm_adb=float(request.POST['coaloading_vm_adb'])
        request.POST['coaloading_vm_arb']=round(coaloading_vm_adb*(100-coaloading_tm_arb)/(100-coaloading_im_adb), 2)

        #rumus coaloading_vm_db
        request.POST['coaloading_vm_db']=round((coaloading_vm_adb*100)/(100-coaloading_im_adb), 2)

        #rumus coaloading_vm_dafb
        coaloading_ash_content_adb=float((request.POST['coaloading_ash_content_adb']))
        request.POST['coaloading_vm_dafb']=round(float(coaloading_vm_adb*100/(100-coaloading_im_adb-coaloading_ash_content_adb)), 2)

        #rumus coaloading_fc_arb
        coaloading_fc_adb=float(request.POST['coaloading_fc_adb'])
        request.POST['coaloading_fc_arb']=round(coaloading_fc_adb*(100-coaloading_tm_arb)/(100-coaloading_im_adb), 2)

        #rumus coaloading_fc_db
        request.POST['coaloading_fc_db']=round((coaloading_fc_adb*100)/(100-coaloading_im_adb), 2)

        #rumus coaloading_fc_dafb
        request.POST['coaloading_fc_dafb']=round((coaloading_fc_adb*100/(100-coaloading_im_adb-coaloading_ash_content_adb)), 2)

        #rumus coaloading_total_sulphur_arb
        coaloading_total_sulphur_adb=float(request.POST['coaloading_total_sulphur_adb'])
        request.POST['coaloading_total_sulphur_arb']=round(coaloading_total_sulphur_adb*(100-coaloading_tm_arb)/(100-coaloading_im_adb), 2)

        #rumus coaloading_total_sulphur_db
        request.POST['coaloading_total_sulphur_db']=round((coaloading_total_sulphur_adb*100)/(100-coaloading_im_adb), 2)

        #rumus coaloading_total_sulphur_dafb
        request.POST['coaloading_total_sulphur_dafb']=round((coaloading_total_sulphur_adb*100/(100-coaloading_im_adb-coaloading_ash_content_adb)), 2)

        #rumus coaunloading_gcv_arb_kcal_kg
        coaunloading_gcv_adb_kcal_kg=float(request.POST['coaunloading_gcv_adb_kcal_kg'])
        coaunloading_tm_arb=float(request.POST['coaunloading_tm_arb'])
        coaunloading_im_adb=float(request.POST['coaunloading_im_adb'])
        request.POST['coaunloading_gcv_arb_kcal_kg']=round(coaunloading_gcv_adb_kcal_kg*(100-coaunloading_tm_arb)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_gcv_db_kcal_kg
        coaunloading_gcv_adb_kcal_kg=float(request.POST['coaunloading_gcv_adb_kcal_kg'])
        coaunloading_im_adb=float(request.POST['coaunloading_im_adb'])
        request.POST['coaunloading_gcv_db_kcal_kg']=round((coaunloading_gcv_adb_kcal_kg*100)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_gcv_dafb_kcal_kg
        coaunloading_ash_content_adb=float((request.POST['coaunloading_ash_content_adb']))
        request.POST['coaunloading_gcv_dafb_kcal_kg']=round(int(coaunloading_gcv_adb_kcal_kg*100/(100-coaunloading_im_adb-coaunloading_ash_content_adb)), 2)

        #rumus coaunloading_ash_content_arb
        request.POST['coaunloading_ash_content_arb']=round(coaunloading_ash_content_adb*(100-coaunloading_tm_arb)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_ash_content_db
        request.POST['coaunloading_ash_content_db']=round((coaunloading_ash_content_adb*100)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_vm_arb
        coaunloading_vm_adb=float(request.POST['coaunloading_vm_adb'])
        request.POST['coaunloading_vm_arb']=round(coaunloading_vm_adb*(100-coaunloading_tm_arb)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_vm_db
        request.POST['coaunloading_vm_db']=round((coaunloading_vm_adb*100)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_vm_dafb
        coaunloading_ash_content_adb=float((request.POST['coaunloading_ash_content_adb']))
        request.POST['coaunloading_vm_dafb']=round(float(coaunloading_vm_adb*100/(100-coaunloading_im_adb-coaunloading_ash_content_adb)), 2)

        #rumus coaunloading_fc_arb
        coaunloading_fc_adb=float(request.POST['coaunloading_fc_adb'])
        request.POST['coaunloading_fc_arb']=round(coaunloading_fc_adb*(100-coaunloading_tm_arb)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_fc_db
        request.POST['coaunloading_fc_db']=round((coaunloading_fc_adb*100)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_fc_dafb
        request.POST['coaunloading_fc_dafb']=round((coaunloading_fc_adb*100/(100-coaunloading_im_adb-coaunloading_ash_content_adb)), 2)

        #rumus coaunloading_total_sulphur_arb
        coaunloading_total_sulphur_adb=float(request.POST['coaunloading_total_sulphur_adb'])
        request.POST['coaunloading_total_sulphur_arb']=round(coaunloading_total_sulphur_adb*(100-coaunloading_tm_arb)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_total_sulphur_db
        request.POST['coaunloading_total_sulphur_db']=round((coaunloading_total_sulphur_adb*100)/(100-coaunloading_im_adb), 2)

        #rumus coaunloading_total_sulphur_dafb
        request.POST['coaunloading_total_sulphur_dafb']=round((coaunloading_total_sulphur_adb*100/(100-coaunloading_im_adb-coaunloading_ash_content_adb)), 2)

        #rumus coaunloading_n_dafb
        coaunloading_n_adb=float(request.POST['coaunloading_n_adb'])
        request.POST['coaunloading_n_dafb']=round((coaunloading_n_adb*100/(100-coaunloading_im_adb-coaunloading_ash_content_adb)), 2)

        #rumus coaunloading_o_adb
        coaunloading_c_adb=float(request.POST['coaunloading_c_adb'])
        coaunloading_h_adb=float(request.POST['coaunloading_h_adb'])
        request.POST['coaunloading_o_adb']=round((100-(coaunloading_im_adb+coaunloading_ash_content_adb+coaunloading_total_sulphur_adb+coaunloading_c_adb+coaunloading_h_adb+coaunloading_n_adb)), 2)

        #rumus labip_gcv_arb_kcal_kg
        labip_gcv_adb_kcal_kg=float(request.POST['labip_gcv_adb_kcal_kg'])
        labip_tm_arb=float(request.POST['labip_tm_arb'])
        labip_im_adb=float(request.POST['labip_im_adb'])
        request.POST['labip_gcv_arb_kcal_kg']=round(labip_gcv_adb_kcal_kg*(100-labip_tm_arb)/(100-labip_im_adb), 2)

        #rumus labip_gcv_db_kcal_kg
        labip_gcv_adb_kcal_kg=float(request.POST['labip_gcv_adb_kcal_kg'])
        labip_im_adb=float(request.POST['labip_im_adb'])
        request.POST['labip_gcv_db_kcal_kg']=round((labip_gcv_adb_kcal_kg*100)/(100-labip_im_adb), 2)

        #rumus labip_gcv_dafb
        labip_ash_content_adb=float((request.POST['labip_ash_content_adb']))
        request.POST['labip_gcv_dafb']=int(round(labip_gcv_adb_kcal_kg*100/(100-labip_im_adb-labip_ash_content_adb), 2))

        #rumus labip_ash_content_arb
        request.POST['labip_ash_content_arb']=round(labip_ash_content_adb*(100-labip_tm_arb)/(100-labip_im_adb), 2)

        #rumus labip_ash_content_db
        request.POST['labip_ash_content_db']=round((labip_ash_content_adb*100)/(100-labip_im_adb), 2)

        #rumus labip_vm_arb
        labip_vm_adb=float(request.POST['labip_vm_adb'])
        request.POST['labip_vm_arb']=round(labip_vm_adb*(100-labip_tm_arb)/(100-labip_im_adb), 2)

        #rumus labip_vm_db
        request.POST['labip_vm_db']=round((labip_vm_adb*100)/(100-labip_im_adb),2)

        #rumus labip_vm_dafb
        labip_ash_content_adb=float((request.POST['labip_ash_content_adb']))
        request.POST['labip_vm_dafb']=round(float(labip_vm_adb*100/(100-labip_im_adb-labip_ash_content_adb)), 2)

        #rumus labip_fc_arb
        labip_fc_adb=float(request.POST['labip_fc_adb'])
        request.POST['labip_fc_arb']=round(labip_fc_adb*(100-labip_tm_arb)/(100-labip_im_adb), 2)

        #rumus labip_fc_db
        request.POST['labip_fc_db']=round((labip_fc_adb*100)/(100-labip_im_adb), 2)

        #rumus labip_fc_dafb
        request.POST['labip_fc_dafb']=round((labip_fc_adb*100/(100-labip_im_adb-labip_ash_content_adb)), 2)

        #rumus labip_total_sulphur_arb
        labip_total_sulphur_adb=float(request.POST['labip_total_sulphur_adb'])
        request.POST['labip_total_sulphur_arb']=round(labip_total_sulphur_adb*(100-labip_tm_arb)/(100-labip_im_adb), 2)

        #rumus labip_total_sulphur_db
        request.POST['labip_total_sulphur_db']=round((labip_total_sulphur_adb*100)/(100-labip_im_adb), 2)

        #rumus labip_total_sulphur_dafb
        request.POST['labip_total_sulphur_dafb']=round((labip_total_sulphur_adb*100/(100-labip_im_adb-labip_ash_content_adb)), 2)

        #rumus labip_n_dafb
        labip_n_adb=float(request.POST['labip_n_adb'])
        request.POST['labip_n_dafb']=round((labip_n_adb*100/(100-labip_im_adb-labip_ash_content_adb)), 2)

        #rumus labip_o_adb
        labip_c_adb=float(request.POST['labip_c_adb'])
        labip_h_adb=float(request.POST['labip_h_adb'])
        #request.POST['labip_o_adb']=(100-(labip_im_adb+labip_ash_content_adb+labip_total_sulphur_adb+labip_c_adb+labip_h_adb+labip_n_adb))

        request.POST._mutable = False

        if request.method == 'POST':
            #form = FormRekap(request.POST)

            instance = get_object_or_404(Rekap_shipment, pk=pk)
            form = FormRekap(request.POST or None, instance=instance)

            if form.is_valid():
                form.save()
            return HttpResponseRedirect('/blog')
        else:
            return HttpResponseRedirect('/blog')
